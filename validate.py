#!/usr/bin/env python3
"""
validate.py — quick sanity check for delegated_authorization_research.xlsx

Run after every edit to build.py. Loads the workbook, prints tab counts,
confirms the Index total equals the row sum, and lint-checks rows and URLs.

Usage:
    python validate.py [path/to/workbook.xlsx]

Default path is ./delegated_authorization_research.xlsx
"""
import sys
import re
from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = sys.argv[1] if len(sys.argv) > 1 else "delegated_authorization_research.xlsx"

CONTENT_TABS = [
    "Published RFCs",
    "Active IETF Drafts",
    "OpenID Foundation",
    "Other Standards & Govt",
    "Academic Papers",
    "Industry & Implementations",
]

# Actual column layout (1-indexed, with header row at row 1):
#   col 1: #               (row number)
#   col 2: Title
#   col 3: One-Sentence Summary
#   col 4: Link
#   col 5: Standards Organization
#   col 6: Comments
TITLE_COL = 2
URL_COL = 4
LAST_COL = 6
URL_PATTERN = re.compile(r"^https?://[^\s]+$")


def main() -> int:
    path = Path(WORKBOOK_PATH)
    if not path.exists():
        print(f"ERROR: workbook not found at {path.resolve()}")
        return 2

    wb = load_workbook(path, data_only=False)  # keep formulas so we can inspect them
    issues: list[str] = []

    # ---- Tab inventory and counts ----
    print(f"Workbook: {path.resolve()}")
    print()
    print(f"  {'Tab':<32} {'Rows':>6}")
    print(f"  {'-' * 32} {'-' * 6}")

    counts: dict[str, int] = {}
    for tab in CONTENT_TABS:
        if tab not in wb.sheetnames:
            issues.append(f"missing tab: {tab}")
            continue
        ws = wb[tab]
        # Header row exists, so data row count is max_row - 1
        n = max(0, ws.max_row - 1)
        counts[tab] = n
        print(f"  {tab:<32} {n:>6}")
    total = sum(counts.values())
    print(f"  {'-' * 32} {'-' * 6}")
    print(f"  {'TOTAL':<32} {total:>6}")
    print()

    # ---- Index total reconciliation ----
    if "Index" in wb.sheetnames:
        idx = wb["Index"]
        # Index C11 holds the TOTAL formula
        index_cell = idx["C11"]
        formula = index_cell.value
        # Reload data_only to get computed value
        wb_calc = load_workbook(path, data_only=True)
        idx_calc = wb_calc["Index"]
        computed = idx_calc["C11"].value
        print(f"Index C11 formula:  {formula}")
        print(f"Index C11 computed: {computed}")
        print(f"Row-sum total:      {total}")
        if computed is None:
            print("  NOTE: computed value is None — workbook needs a recalc pass.")
            print("        Open it once in Excel/LibreOffice, or run a recalc tool,")
            print("        so the SUM formula populates. Then re-run validate.py.")
        elif computed != total:
            issues.append(f"Index C11 ({computed}) != row sum ({total})")
        else:
            print("  ✓ Index total matches row sum.")
    else:
        issues.append("missing Index tab")
    print()

    # ---- Row-level lint: title presence, URL well-formedness ----
    print("Row lint:")
    lint_count = 0
    for tab in CONTENT_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        # Data starts at row 2 (row 1 is header).
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, LAST_COL + 1)]
            title = row[TITLE_COL - 1]
            url = row[URL_COL - 1]
            # Skip fully empty rows
            if all(v in (None, "") for v in row):
                continue
            if not title:
                issues.append(f"{tab} row {r}: missing title")
                lint_count += 1
            if not url:
                issues.append(f"{tab} row {r}: missing URL")
                lint_count += 1
            elif not URL_PATTERN.match(str(url)):
                issues.append(f"{tab} row {r}: malformed URL: {str(url)[:60]}")
                lint_count += 1
    if lint_count == 0:
        print("  ✓ No row-level issues.")
    print()

    # ---- Summary ----
    if issues:
        print(f"FOUND {len(issues)} ISSUE(S):")
        for issue in issues:
            print(f"  - {issue}")
        return 1
    else:
        print("All checks passed.")
        return 0


if __name__ == "__main__":
    sys.exit(main())
