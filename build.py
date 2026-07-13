from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
wb.remove(wb.active)

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
BODY_FONT   = Font(name='Arial', size=10)
LINK_FONT   = Font(name='Arial', size=10, color='0563C1', underline='single')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_ALIGN   = Alignment(vertical='top', wrap_text=True)
THIN = Side(border_style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLORS = {
    'RFC':       '1F4E78',
    'Drafts':    'C0504D',
    'OpenID':    '7030A0',
    'Other':     '548235',
    'Academic':  'BF8F00',
    'Industry':  '404040',
    'Summary':   '305496',
}

COLS = ["#", "Title", "One-Sentence Summary", "Link", "Standards Organization", "Comments"]
WIDTHS = {1: 5, 2: 44, 3: 70, 4: 56, 5: 26, 6: 62}

def make_sheet(title, color, rows):
    ws = wb.create_sheet(title)
    fill = PatternFill('solid', start_color=color)
    for col, h in enumerate(COLS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, fill, HEADER_ALIGN, BORDER
    for i, (t, s, link, org, comment) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=t)
        ws.cell(row=i, column=3, value=s)
        link_cell = ws.cell(row=i, column=4, value=link)
        ws.cell(row=i, column=5, value=org)
        ws.cell(row=i, column=6, value=comment)
        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            cell.font = BODY_FONT
        link_cell.hyperlink = link
        link_cell.font = LINK_FONT
    for col, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    return ws

# ============================================================
# TAB 1: PUBLISHED RFCs
# ============================================================
rfc_rows = [
    ("RFC 9635 — Grant Negotiation and Authorization Protocol (GNAP)",
     "An IETF standards-track RFC for a next-generation delegation protocol that removes pre-registration by letting clients present a key on first contact with the AS.",
     "https://www.rfc-editor.org/rfc/rfc9635",
     "IETF (concluded GNAP WG)",
     "Often called 'OAuth 3'; published Oct 2024, working group concluded so the spec is in maintenance mode."),

    ("RFC 9396 — OAuth 2.0 Rich Authorization Requests (RAR)",
     "An IETF RFC introducing the authorization_details JSON parameter so OAuth clients can request structured fine-grained permissions instead of flat scopes.",
     "https://datatracker.ietf.org/doc/html/rfc9396",
     "IETF",
     "Foundational to FAPI 2.0, open-banking PSD2/PSD3, and increasingly used as the policy carrier for MCP and agent transaction tokens."),

    ("RFC 8693 — OAuth 2.0 Token Exchange",
     "An IETF RFC defining a token-exchange grant with subject_token and actor_token parameters, providing the canonical mechanism for impersonation and delegation in OAuth.",
     "https://www.rfc-editor.org/rfc/rfc8693",
     "IETF",
     "The 'actor_token' chain is the basis for almost every multi-hop agent-delegation proposal but is itself subject to the 'delegation chain splicing' attack."),

    ("RFC 9700 — Best Current Practice for OAuth 2.0 Security",
     "The IETF BCP that codifies current OAuth security requirements such as PKCE, exact redirect matching, and sender-constrained tokens.",
     "https://www.rfc-editor.org/rfc/rfc9700",
     "IETF",
     "Published January 2025; companion to OAuth 2.1 — implementations claiming OAuth 2.1 conformance are essentially asserting RFC 9700 conformance."),
]
make_sheet("Published RFCs", COLORS['RFC'], rfc_rows)

# ============================================================
# TAB 2: ACTIVE IETF DRAFTS
# ============================================================
draft_rows = [
    # ---- IETF Working Group Charters (scope-defining documents) ----
    ("Charter: OAuth WG Proposed Recharter (charter-ietf-oauth-05-05)",
     "The OAuth WG's proposed recharter (currently under External Review, on the 2026-06-04 IESG telechat agenda) that formally adds 'Complex Delegation' to the work program — new mechanisms and extensions for authorization of automated agents acting on behalf of users, including cross-administrative-domain scenarios.",
     "https://datatracker.ietf.org/doc/charter-ietf-oauth/05-05/",
     "IETF (OAuth WG charter)",
     "★ Hugely significant: the OAuth WG is formally re-chartering around agent delegation. Milestones include Transaction Tokens and OAuth 2.1 to IESG by Dec 2026, SD-JWT VC by Jul 2026. Coordinates with WIMSE on multi-hop workload identity."),

    ("Charter: Web Bot Auth WG (charter-ietf-webbotauth-01)",
     "The formal IETF charter for the newly-approved Web Bot Auth WG, standardizing methods for bots (search crawlers, AI training crawlers, AI agents retrieving content for end users) to cryptographically authenticate themselves to websites built for humans.",
     "https://datatracker.ietf.org/doc/charter-ietf-webbotauth/01/",
     "IETF (WebBotAuth WG charter)",
     "Approved Oct 2025; complements but does not duplicate OAuth delegation work — authenticates the bot/agent itself, not the end user. Liaises with AIPREF, HTTPBIS, OAuth, TLS, and WIMSE."),

    ("Charter: AI Preferences WG (charter-ietf-aipref-01)",
     "The formal IETF charter for the AIPREF WG, standardizing vocabulary and protocol mechanisms (e.g., extending the Robots Exclusion Protocol and HTTP headers) for content owners to express preferences about AI training, deployment, and use of their content.",
     "https://datatracker.ietf.org/doc/charter-ietf-aipref/01/",
     "IETF (AIPREF WG charter)",
     "Approved April 2025; the charter explicitly puts 'authenticating or authorizing clients and/or crawlers' out of scope — included here for ecosystem context, not because it's delegated-authz work per se. Sibling WG to WebBotAuth."),

    ("draft-king-dawn-requirements — Requirements for Discovery of Agents, Workloads, and Named Entities (DAWN)",
     "An individual IETF draft from Adrian Farrel and Daniel King setting out solution-neutral requirements for discovering AI agents, services, and workloads across administrative boundaries — what must be discoverable, what trust properties apply, and what architectural constraints exist.",
     "https://datatracker.ietf.org/doc/draft-king-dawn-requirements/01/",
     "IETF (individual)",
     "Revision -01, April 2026; intentionally NOT a protocol — sits one layer below auth. Authentication and authorization of discovered entities are out of scope, but DAWN discovery happens *before* auth/authz can begin."),

    # ---- WG drafts (more weight, closer to standardization) ----
    ("draft-ietf-oauth-v2-1 — The OAuth 2.1 Authorization Framework",
     "An IETF WG draft that obsoletes RFC 6749/6750 by mandating PKCE for the authorization code grant, removing implicit and ROPC, and requiring exact redirect-URI matching.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-v2-1/",
     "IETF (OAuth WG)",
     "Revision -15 (March 2026) is the current baseline that MCP, FAPI 2.0, and most agent specs profile against."),

    ("draft-ietf-oauth-identity-chaining — OAuth Identity and Authorization Chaining Across Domains",
     "An IETF WG draft that defines how to preserve identity and authorization context across trust domains by combining RFC 8693 token exchange with RFC 7521/7523 JWT assertions.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-identity-chaining/",
     "IETF (OAuth WG)",
     "IESG-approved; currently in RFC Editor queue (status: Waiting on Authors — outstanding SECDIR 'Has issues' and OPSDIR 'Not ready' flags unresolved; no RFC number assigned yet as of Jul 2026). The canonical multi-domain delegation pattern referenced by most agent and zero-trust drafts."),

    ("draft-ietf-oauth-identity-assertion-authz-grant — Identity Assertion JWT Authorization Grant (ID-JAG)",
     "An IETF WG draft (Parecki/McGuinness/Campbell) defining how an app uses an identity assertion to obtain an access token for a third-party API by coordinating through a shared enterprise IdP.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-identity-assertion-authz-grant/",
     "IETF (OAuth WG)",
     "Revision -03, April 2026; the cross-IdP SSO-to-API bridge that the McGuinness Actor Profile draft layers on top of, and that WorkOS's auth.md uses as one of its three discovery flows."),

    ("draft-ietf-oauth-transaction-tokens — Transaction Tokens (Txn-Tokens)",
     "An IETF WG draft defining short-lived signed JWTs that propagate immutable user identity and authorization context through internal call chains within a trust domain.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-transaction-tokens/",
     "IETF (OAuth WG)",
     "Revision -09 published 6 Jul 2026 (was -08 Mar 2026); co-authored by Tulshibagwale (CAEP inventor), Fletcher, and Kasselman. -09 reorganized the Request Context section, changed JWT body claims from OPTIONAL to RECOMMENDED, and enhanced security considerations for invalidated access tokens. Still in WGLC; IESG submission milestone Dec 2026."),

    ("draft-ietf-oauth-first-party-apps — OAuth 2.0 for First-Party Applications",
     "An IETF WG draft defining an Authorization Challenge Endpoint that lets first-party native apps drive a browserless OAuth flow while still supporting step-up authentication via RFC 9470.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-first-party-apps/",
     "IETF (OAuth WG)",
     "Revision -04, published 1 Jul 2026; explicitly excluded for third-party use because it requires high AS↔client trust."),

    ("draft-ietf-oauth-attestation-based-client-auth — OAuth 2.0 Attestation-Based Client Authentication",
     "An IETF WG draft (Looker/Bastian/Bormann) introducing two JWTs — a Client Attestation issued by a Client Attester and a Client Attestation PoP signed by the client instance — that travel in HTTP headers to let traditionally-public clients authenticate to the AS without a shared secret.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-attestation-based-client-auth/",
     "IETF (OAuth WG)",
     "Revision -10, Jul 2026 (was -08 Mar 2026). -10 added a 'client_attestation_pop_methods_supported' metadata parameter, created a new 'OAuth Client Attestation Proof-of-Possession Methods' registry, clarified that PoP mechanisms from other specs are permitted, and refined DPoP combined-mode handling. Relates to the RATS Passport Model per RFC 9334; RATS attestation procedures themselves are deliberately out of scope. The 'client instance' framing pairs naturally with the McGuinness Client Instance Assertion and AI Agent Instance drafts."),

    ("draft-ietf-oauth-spiffe-client-auth — OAuth SPIFFE Client Authentication",
     "An IETF WG draft (Schwenkschuster/Kasselman/Rose-NIST/Thorgersen-IBM) profiling RFC 7521, RFC 7523, and OAuth attestation-based client auth so that SPIFFE-enabled workloads can authenticate to OAuth ASes using JWT-SVID, WIT-SVID, or X.509-SVID credentials instead of shared client secrets.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-spiffe-client-auth/",
     "IETF (OAuth WG)",
     "Revision -01, March 2026 (re-slugged from draft-schwenkschuster-oauth-spiffe-client-auth after WG adoption); the protocol bridge between SPIFFE/WIMSE workload identity and the OAuth client-auth surface."),

    ("draft-kahrer-oauth-client-challenge-protocol — OAuth Client Challenge Protocol",
     "An IETF individual draft from Judith Kahrer (Curity) defining a new 'insufficient_client_authorization' error code so the authorization server can dynamically challenge the client mid-flow for an additional assertion, verifiable presentation, or proof-of-possession material — enabling just-in-time and step-up client authorization.",
     "https://datatracker.ietf.org/doc/draft-kahrer-oauth-client-challenge-protocol/",
     "IETF (individual)",
     "Revision -00 published 19 May 2026; deliberately distinct from first-party-apps (challenges the CLIENT, not the user) and works for both first- and third-party clients. Cites Mastercard Verifiable Intent as a use case, linking it directly to the mandate/RAR pattern OVID implements."),

    ("draft-ietf-oauth-browser-based-apps — OAuth 2.0 for Browser-Based Applications",
     "An IETF WG draft establishing the Backend-for-Frontend (BFF) and other architectural patterns as current best practice for SPAs given browser-resident token-storage limitations.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-browser-based-apps/",
     "IETF (OAuth WG)",
     "Revision -26 (Dec 2025); critical for any delegated-authorization design touching SPAs because it codifies why pure-browser refresh tokens are deprecated."),

    ("draft-ietf-oauth-security-topics-update — Updates to OAuth 2.0 Security BCP",
     "An IETF WG draft that extends RFC 9700 with new countermeasures, notably rules against audience-injection attacks where a client authenticates to multiple authorization servers.",
     "https://datatracker.ietf.org/doc/draft-ietf-oauth-security-topics-update/",
     "IETF (OAuth WG)",
     "Revision -01, March 2026; co-authored by University of Stuttgart (FAPI formal-analysis team) so the recommendations are formally grounded."),

    ("draft-mcguinness-oauth-actor-profile — OAuth Actor Profile for Delegation",
     "Karl McGuinness's individual draft profiling the RFC 8693 'act' claim with required iss, optional sub_profile entity classification, and uniform processing across JWT assertion grants, JWT access tokens, and Transaction Tokens.",
     "https://datatracker.ietf.org/doc/html/draft-mcguinness-oauth-actor-profile-00",
     "IETF (individual)",
     "Revision -00 published 30 April 2026; defines depth-1+ chain validation, cycle-aware construction, and bearer-to-PoP upgrade — the most complete actor-representation profile to date."),

    ("draft-mw-oauth-actor-chain — Cryptographically Verifiable Actor Chains for OAuth 2.0 Token Exchange",
     "An IETF individual draft (Prasad/Krishnan/Lopez/Addepalli) defining six actor-chain profiles for RFC 8693 — Declared/Verified × Full/Subset/Actor-Only Disclosure — that preserve and cryptographically validate delegation-path continuity across successive token exchanges instead of relying on RFC 8693's informational-only nested act claims.",
     "https://datatracker.ietf.org/doc/draft-mw-oauth-actor-chain/",
     "IETF (individual)",
     "Revision -01, 30 Jun 2026 (was -00 May 2026; re-slugged from draft-mw-spice-actor-chain-05); a complementary alternative to the McGuinness Actor Profile that directly addresses RFC 8693's chain-splicing weakness by making the chain itself cryptographically verifiable rather than just informational."),

    ("draft-mcguinness-oauth-client-instance-assertion — OAuth Client Instance Assertion",
     "Karl McGuinness's individual draft for representing a specific running instance of an OAuth client (rather than just the client registration) so authorization servers can bind grants to a concrete instance.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-client-instance-assertion/",
     "IETF (individual)",
     "Note: full draft text was not retrievable through automated fetch at time of writing — recommend reading directly. Companion to the Actor Profile draft and to attestation-based client auth."),

    ("draft-mcguinness-oauth-resource-token-resp — OAuth 2.0 Resource Parameter in Access Token Response",
     "A McGuinness/Skokan individual draft adding a 'resource' parameter to the OAuth access token response so clients can confirm which resource the issued token is valid for, mitigating resource mix-up attacks especially in deployments that use RFC 8707 Resource Indicators.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-resource-token-resp/",
     "IETF (individual)",
     "Revision -03 (March 2026). Filip Skokan added as co-author in -03. Complements RFC 8707 / RFC 9700 / RFC 9207 by adding issuance-time confirmation to the discovery-time mechanisms. Notable for the agent-authz corpus because dynamic resource discovery is a core agent pattern and resource-mix-up attacks are amplified when agents juggle many short-lived tokens."),

    ("draft-mcguinness-oauth-rfc9728bis — Update to OAuth 2.0 Protected Resource Metadata Resource Identifier Validation",
     "A McGuinness individual draft updating Section 3.3 of RFC 9728 (Protected Resource Metadata) so the resource value can be any URI sharing the same TLS origin as the requested URL whose path is a prefix of it — broadening the set of WWW-Authenticate response use cases without changing the rest of RFC 9728.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-rfc9728bis/",
     "IETF (individual)",
     "Revision -01 (Feb 2026). A small targeted correction to the PRM validation rule. Notable as supporting infrastructure for the McGuinness Mission-Bound OAuth work — PRM is the substrate for Resource AS / Resource Server discovery of required RAR types under that profile."),

    ("draft-mcguinness-token-xchg-target-svc-disco — OAuth 2.0 Token Exchange Target Service Discovery",
     "A McGuinness/Parecki individual draft defining a discovery endpoint that, given a subject token, returns the set of available target services (audiences, resources, scopes) the holder can request via RFC 8693 Token Exchange — accepting any registered subject_token_type so it supports advanced flows including identity chaining and cross-domain delegation.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-token-xchg-target-svc-disco/",
     "IETF (individual)",
     "Revision -01 (Feb 2026), co-authored with Aaron Parecki (Okta). Fills a discovery gap that ID-JAG and the Transaction Token Chaining Profile both implicitly assume — clients today must know which audience to request, which doesn't compose with open-world agentic flows where the right downstream service is determined at runtime."),

    ("draft-mcguinness-oauth-actor-receipts — OAuth Actor Receipts for Delegation Provenance",
     "A McGuinness companion draft to the Actor Profile, aiming to provide per-hop signed receipts as the verifiable provenance layer that the Actor Profile (which keeps the 'act' chain informational) deliberately leaves out. Pulls over the provenance gap from the core profile so each principal in a delegation chain can produce independent cryptographic evidence.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-actor-receipts/",
     "IETF (individual)",
     "Revision -00, 4 Jul 2026; now formally on IETF Datatracker (was GitHub-only pre-publication through Jun 2026). Recommended companion to the Actor Profile when a deployment needs Verified Full Disclosure mode (per the Mission-Bound OAuth Runtime Enforcement Profile's Actor Provenance Module). Completes a triad with actor-profile and actor-proofs — receipts provide provenance, proofs provide cryptographic per-hop accountability."),

    ("draft-mcguinness-oauth-actor-proofs — OAuth Actor-Signed Hop Proofs",
     "Introduces the 'actor_proofs' claim — a signed per-hop proof chain where each actor signs its own participation and the target binding it authorized for that hop; proofs are hash-chained and validated using actor verification keys from trusted sources, with optional references to sibling actor receipts.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-actor-proofs/",
     "IETF (individual)",
     "Revision -00, 4 Jul 2026. Completes the actor triad with actor-profile (chain structure) and actor-receipts (provenance) — actor-proofs closes the cryptographic accountability gap. Filed the same day as ai-agent-instance, id-assertion-framework, and domain-authorized-issuer — McGuinness submitted four new I-Ds on Jul 4-6, bringing his total to 11 individual I-Ds on Datatracker."),

    ("draft-mcguinness-oauth-insufficient-claims — OAuth 2.0 Insufficient Claims Challenge",
     "A McGuinness individual draft defining an insufficient_claims error code and required_claims parameter enabling authorization servers and protected resources to signal precisely which claims are missing from a presented credential. Allows resource servers to request enriched credentials carrying specific missing attributes rather than returning a generic 401.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-insufficient-claims/",
     "IETF (individual)",
     "Revision -00, 27 May 2026. Enables just-in-time claim negotiation in multi-agent delegation chains where the acting party must assemble claims from multiple sources. Natural companion to the Actor Profile and Client Instance Assertion — those drafts define the identity structure; this one provides the feedback loop when that structure is incomplete at the RS."),

    ("draft-mcguinness-oauth-ai-agent-instance — OAuth 2.0 AI Agent Instance Profile",
     "Profiles OAuth 2.0 for deployments where a single client ID represents an agent platform running many concurrent agent instances; defines claims conveying attested agent instance identity and provenance from an agent attester to the AS, with delegation-chain semantics for sub-agents. Claims are carrier-independent, compatible with both Client Instance Assertions and Client Attestation JWT.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-ai-agent-instance/",
     "IETF (individual)",
     "Revision -00, 4 Jul 2026. Fills the gap between a registered OAuth client ID (platform-level) and a running agent instance (runtime-level) — the level at which delegation chains, attestation, and revocation actually operate. Companion to draft-mcguinness-oauth-client-instance-assertion (general instance framing) and draft-mcguinness-oauth-actor-profile (chain representation)."),

    ("draft-mcguinness-oauth-id-assertion-framework — OAuth Identity Assertion Trust Framework",
     "Addresses the gap where issuer authentication alone does not prove AS authority over subject namespaces; defines an Authority Delegation Model with independent trust-evaluation categories and an Identity Assertion Issuer Trust Policy (JSON document declaring required trust methods) that Resource ASes evaluate before accepting identity assertions.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-id-assertion-framework/",
     "IETF (individual)",
     "Revision -00, 4 Jul 2026. Directly motivates and companions draft-mcguinness-oauth-domain-authorized-issuer (below) which provides a DNS-based trust method. Addresses the 'who is allowed to assert identities in this namespace?' question that ID-JAG and the identity-chaining WG draft assume resolved but do not specify."),

    ("draft-mcguinness-oauth-domain-authorized-issuer — OAuth Domain-Authorized Issuer Trust Method",
     "DNS domain owners publish a policy (DNS TXT record at _oauth-issuer-policy.{domain} or HTTPS well-known fallback) listing the OAuth ASes authorized to assert identities in that namespace; Resource ASes use this to validate assertion issuers before accepting identity assertions or chaining tokens. Pattern echoes CAA/SPF/DKIM. Non-transitive — flat authorization, no delegation chains.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-domain-authorized-issuer/",
     "IETF (individual)",
     "Revision -00, 4 Jul 2026. Implements one trust method for draft-mcguinness-oauth-id-assertion-framework. DNS-based approach is lightweight and operationally familiar; the non-transitive design avoids the amplification risks of transitive trust chains. Closes the 'how do you know the AS is allowed to speak for this domain?' gap that cross-domain identity chaining has always assumed away."),

    ("draft-mcguinness-oauth-mission — Mission-Bound Authorization for OAuth 2.0",
     "Introduces 'Mission' — a durable, integrity-bound artifact stored at the AS representing an approved task; clients submit Mission Intent via PAR, ASes derive concrete permissions, and approvers grant consent via integrity anchors. Access tokens carry a 'mission' claim; issuance is gated by the Mission's lifecycle state, enabling revocation governance and audit trails for multi-step agent operations.",
     "https://datatracker.ietf.org/doc/draft-mcguinness-oauth-mission/",
     "IETF (individual)",
     "Revision -00, 6 Jul 2026. This is the formal I-D target for the 'Mission-Bound OAuth MVP' blog post (May 2026, Industry tab) — the five-wire-addition protocol that makes the user-approved task a first-class OAuth object. Paired with the runtime enforcement profile (still not on Datatracker as of Jul 10). The Mission lifecycle (pending_approval → active → suspended → revoked → expired → completed → rejected) and the proposal_hash/consent_rendering_hash binding anchor what was approved versus what the user saw."),

    ("draft-mishra-oauth-agent-grants — Delegated Agent Authorization Protocol (DAAP)",
     "An IETF individual draft addressing runtime agent identity for dynamically-spawned agents, multi-agent sub-delegation with depth-limiting, and cascade revocation of an entire delegation subtree when any ancestor is revoked.",
     "https://datatracker.ietf.org/doc/draft-mishra-oauth-agent-grants/",
     "IETF (individual)",
     "Revision -01 (March 2026); registers new JWT claims (agt, dev, grnt, scp, bdg) — a more aggressive alternative to RFC 8693 with budget-bounded grants. Overlaps with draft-niyikiza-oauth-attenuating-agent-tokens; OAuth WG will likely expect consolidation."),

    ("draft-kuehlewind-audit-architecture — Architecture for Auditing AI Agent Delegation and Interactions",
     "An IETF individual draft from Mirja Kühlewind (Ericsson, former IETF TSV-AD) and Henk Birkholz (Fraunhofer SIT, lead author of RFC 9334 RATS and SCITT architecture) defining four record types — Interaction, Action, Delegation, and Authorization Transition — and an Audit Context that propagates across OAuth + WIMSE + RATS + SCITT to make agent behaviour verifiably auditable end-to-end.",
     "https://datatracker.ietf.org/doc/draft-kuehlewind-audit-architecture/",
     "IETF (individual)",
     "Revision -00 published 18 May 2026; the cross-layer integration piece — explicitly composes the McGuinness Actor Profile, Transaction Tokens, identity-chaining, ID-JAG, WIMSE, RATS, and SCITT into a single auditable architecture with 11 proposed work items. The treats-Agent-as-untrusted threat model is what regulators (SOX, PSD2, EU AI Act) actually want."),

    ("draft-birkholz-verifiable-agent-conversations — Verifiable Agent Conversation Records",
     "An IETF individual draft (Birkholz/Heldt/Steele) defining a CDDL data format — with both JSON and CBOR encodings — for tamper-evident, verifiably-signed records of agent conversations, agent reasoning traces, tool calls, and system events that support long-term evidentiary value across chains of custody.",
     "https://datatracker.ietf.org/doc/draft-birkholz-verifiable-agent-conversations/",
     "IETF (individual)",
     "Revision -00 published 25 Feb 2026; presented at IETF 125 dispatch and hotrfc. The canonical Interaction-Record format for the Kuehlewind audit architecture. Motivating examples are concrete: agents acting beyond scope, CoT divergence, and 'sandbagging' during evaluations."),

    ("draft-klrc-aiagent-auth — AI Agent Authentication and Authorization",
     "An IETF individual draft that does not invent new protocols but composes WIMSE/SPIFFE workload identity, OAuth 2.x, and OpenID SSF to authenticate and delegate authority to AI agents.",
     "https://datatracker.ietf.org/doc/html/draft-klrc-aiagent-auth-00",
     "IETF (individual)",
     "Revision -00, March 2026; introduces 'Agent Identity Management System (AIMS)' and is the cleanest standards-based framing of agent-as-workload."),

    # ---- Individual drafts: Identity / Authorization / Delegation cluster ----
    ("draft-niyikiza-oauth-attenuating-agent-tokens — Attenuating Authorization Tokens for Agentic Delegation Chains",
     "An IETF individual draft extending RFC 9396 Rich Authorization Requests with monotonic attenuation at each delegation hop, offline chain verification, and a typed constraint vocabulary for tool-level argument restrictions.",
     "https://datatracker.ietf.org/doc/draft-niyikiza-oauth-attenuating-agent-tokens/",
     "IETF (individual)",
     "Revision -00, March 2026; arguably the strongest purely-OAuth-grounded delegation draft — offline verification is a key scalability property. Good path to OAuth WG adoption; overlaps with draft-mishra-oauth-agent-grants and will likely need to be consolidated with it."),

    ("draft-singla-agent-identity-protocol — AIP: Decentralized Identity and Delegation for AI Agents",
     "An IETF individual draft proposing a six-layer model (identity through reputation) for AI agents with W3C DID anchoring, a Delegation Depth counter, three Grant ceremony tiers (G1/G2/G3), DPoP proof-of-possession, and an MCP integration layer.",
     "https://datatracker.ietf.org/doc/draft-singla-agent-identity-protocol/",
     "IETF (individual)",
     "Revision -00, 17 April 2026; the most comprehensive identity framework in the set. The W3C DID dependency will face IETF pushback, the 75+ page scope needs reduction, and the 'AIP' name collision with two other drafts (draft-prakash-aip, draft-aip-agent-identity-protocol) must be resolved."),

    ("draft-prakash-aip — AIP: Verifiable Delegation for AI Agent Systems",
     "An IETF individual draft introducing Invocation-Bound Capability Tokens (IBCTs) in two modes — compact (JWT/Ed25519) for single-hop and chained (Biscuit + Datalog) for multi-hop — with bindings for MCP, A2A, and HTTP APIs.",
     "https://datatracker.ietf.org/doc/draft-prakash-aip/",
     "IETF (individual)",
     "Revision -00, 27 March 2026; technically strong — Biscuit + Datalog is well-suited to scope attenuation. The 'AIP' name collision with draft-singla-agent-identity-protocol and draft-aip-agent-identity-protocol is a blocking issue for WG adoption."),

    ("draft-goswami-agentic-jwt — Secure Intent Protocol: JWT-Compatible Agentic Identity and Workflow Management",
     "An IETF individual draft defining an Agentic JWT extension to OAuth 2.0 with a Supervisor Agent role for sub-agent coordination, explicitly addressing the 'intent-execution separation' gap and adding workflow-step authorization.",
     "https://datatracker.ietf.org/doc/draft-goswami-agentic-jwt/",
     "IETF (individual)",
     "Revision -00, December 2025; the intent-execution-separation framing is insightful and distinguishing. The workflow-step authorization approach is novel but may face skepticism in the OAuth WG."),

    ("draft-ni-wimse-ai-agent-identity — WIMSE Applicability for AI Agents",
     "An IETF individual draft mapping the WIMSE workload-identity architecture to AI-agent use cases and establishing credential-management mechanisms for agentic AI within the WIMSE framework.",
     "https://datatracker.ietf.org/doc/draft-ni-wimse-ai-agent-identity/",
     "IETF (individual)",
     "Revision -02, February 2026; an important bridge between the WIMSE WG and the AI agent space. If an AI-agent WG is ever chartered, this applicability statement is essential groundwork. Co-authored with Liu, who also leads the related security-requirements draft."),

    ("draft-sharif-agent-identity-framework — Agent Identity Framework: Trust and Identity for Autonomous AI Agents",
     "An IETF individual draft proposing a five-layer model (identity, authorization, attestation, evidence, trust) for autonomous AI agents, with a gap analysis between current Internet standards and agent requirements.",
     "https://datatracker.ietf.org/doc/draft-sharif-agent-identity-framework/",
     "IETF (individual)",
     "Revision -00, 6 April 2026; more gap-analysis than protocol spec. Useful as problem statement. Five-layer model is a reasonable decomposition but insufficient as a standalone contribution. Part of the Sharif cluster (with -agent-audit-trail, -attp, etc.)."),

    ("draft-aip-agent-identity-protocol — AIP: Agentic Authentication and Authorized Policy Enforcement",
     "An IETF individual draft defining an AIP Token signed per tool call (agent ID, tool, nonce, timestamp, signature), an AgentPolicy YAML format for per-tool argument constraints and DLP rules, and a Human-in-the-Loop proxy enforcement model.",
     "https://datatracker.ietf.org/doc/draft-aip-agent-identity-protocol/",
     "IETF (individual)",
     "Revision -00, 16 March 2026; implementation-focused. YAML policy approach is pragmatic but not interoperable-by-design. HITL integration is a valuable safety feature. Third member of the 'AIP' name-collision triple — must be resolved before any can progress."),

    ("draft-liu-oauth-chain-delegation — Delegation Chain for OAuth 2.0",
     "Introduces a delegation_chain JWT claim as a structured companion to RFC 8693's act claim; delegation_chain records the complete delegation history as an ordered array with per-hop authorization constraints and optional cryptographic confirmation from each delegating agent. Supports cross-domain delegation with integrated user consent interaction.",
     "https://datatracker.ietf.org/doc/draft-liu-oauth-chain-delegation/",
     "IETF (individual)",
     "Revision -00, 6 Jun 2026; authors: Dapeng Liu, Judy Zhu, Suresh Krishnan (Ericsson), Aaron Parecki (Okta). Provides the structured delegation lineage that act alone cannot express; strong overlap with draft-mw-oauth-actor-chain and draft-niyikiza-oauth-attenuating-agent-tokens — OAuth WG will likely expect consolidation across these three."),

    ("draft-fletcher-transaction-token-chaining-profile — Transaction Token Authorization Grant Profile for OAuth Identity and Authorization Chaining",
     "Establishes a framework for using Transaction Tokens as subject credentials in OAuth 2.0 Token Exchange requests (RFC 8693), enabling services within one trust domain to obtain JWT authorization grants for accessing resources across domain boundaries without exposing internal token formats.",
     "https://datatracker.ietf.org/doc/draft-fletcher-transaction-token-chaining-profile/",
     "IETF (individual)",
     "Revision -02, 6 Jul 2026 (was -01 Jun 2026); authors: George Fletcher (Practical Identity LLC), Pieter Kasselman (Defakto Security), Sean O'Dell (CVS Health). Directly bridges draft-ietf-oauth-transaction-tokens and draft-ietf-oauth-identity-chaining (both in corpus), closing the 'how do TX Tokens actually cross domains' gap. Note: lead author is the corpus maintainer."),

    ("draft-zhu-oauth-async-delegation — Sender-Constrained Delegation Handle for Asynchronous OAuth 2.0 Identity Chaining",
     "Introduces a Delegation Handle — a sender-constrained, audience-locked JWT issued by authorization servers — enabling acting clients to refresh chained access tokens without re-prompting an offline end user while maintaining strict policy bounds and identity chains.",
     "https://datatracker.ietf.org/doc/draft-zhu-oauth-async-delegation/",
     "IETF (individual)",
     "Revision -00, 22 May 2026; authors: Larry Zhu, Sam Currie (Atlassian). Addresses a concrete gap in agent delegation: what happens when the delegating human is not present to re-authorize. Directly relevant to long-running agentic workflows using draft-ietf-oauth-identity-chaining."),

    ("draft-chen-oauth-agent-revocation — OAuth 2.0 Agent Authorization Explicit Revocation",
     "Extends RFC 7009 token revocation for agent-based scenarios: introduces batch revocation by agent ID, cascading revocation across delegation chains, conditional revocation options, and verifiable audit trails.",
     "https://datatracker.ietf.org/doc/draft-chen-oauth-agent-revocation/",
     "IETF (individual)",
     "Revision -00, 27 Apr 2026; authors: Meiling Chen, Li Su (China Mobile). Fills the lifecycle management gap — existing RFC 7009 revocation is single-token and client-centric, but agent delegation chains require cascade semantics. Natural companion to draft-ietf-oauth-transaction-tokens and draft-niyikiza-oauth-attenuating-agent-tokens."),

    ("draft-jiang-oauth-intent-admission — Intent Admission Assertions for Agentic Systems",
     "Defines a signed, verifiable Intent Admission Assertion (IAA) that an admission point creates after verifying agent identity, evaluating permissions against policy, and — when required — obtaining explicit user consent; the IAA is transmitted to the execution endpoint which re-validates before proceeding. Leverages OAuth and RFC 9396 Rich Authorization Requests.",
     "https://datatracker.ietf.org/doc/draft-jiang-oauth-intent-admission/",
     "IETF (individual)",
     "Revision -00, 23 Jun 2026; authors: Yuning Jiang, Lun Li, Yurong Song, Faye Liu (Huawei). The required_consent path maps directly to the HITL pattern central to the corpus. Companion to draft-jiang-intent-security (threat model) from the same Huawei group."),

    ("draft-ni-oauth-batch-authorization-delegation — Batch Authorization Delegation",
     "Defines a mechanism for delegating a batch of fine-grained, actor-bound permissions in a single request across multiple collaborating actors; uses RFC 9396 RAR to carry per-actor authorization_details and RFC 8693 Token Exchange for sub-agent delegation, targeting multi-agent orchestration where a leader-agent receives batch permissions and delegates subsets to sub-agents.",
     "https://datatracker.ietf.org/doc/draft-ni-oauth-batch-authorization-delegation/",
     "IETF (individual)",
     "Revision -00, 3 Jul 2026; authors: Ni Yuan, Peter Chunchi Liu (Huawei). Addresses the round-trip overhead problem in large-scale multi-agent orchestration — individual delegation exchanges per sub-agent don't scale. Complements draft-song-oauth-ai-agent-collaborate-authz (same Huawei group, coordination focus) and draft-niyikiza-oauth-attenuating-agent-tokens (attenuation semantics). The RAR-based batch approach aligns well with the OAuth WG's RAR investment."),

    ("draft-liu-oauth-a2a-profile — Agent-to-Agent (A2A) Profile for OAuth Transaction Tokens",
     "Defines a profile for using OAuth Transaction Tokens in distributed agent-to-agent communication scenarios; specifies mechanisms for embedding call-chain context within tokens to preserve agent identity, authorization information, and operational flow across agent workloads in trusted environments.",
     "https://datatracker.ietf.org/doc/draft-liu-oauth-a2a-profile/",
     "IETF (individual)",
     "EXPIRED. Revision -00, Oct 2025; authors: Peter Chunchi Liu, Ni Yuan (Huawei) — same author pair as draft-ni-oauth-batch-authorization-delegation. Directly profiles draft-ietf-oauth-transaction-tokens for A2A use; overlaps significantly with draft-araut-oauth-transaction-tokens-for-agents. Never revised; likely deferred rather than abandoned given both authors remain active in the corpus through Jul 2026."),

    ("draft-song-oauth-ai-agent-collaborate-authz — OAuth 2.0 Extension for Multi-AI Agent Collaboration",
     "Extends OAuth 2.0 for coordinated multi-agent task groups: defines a collaborative authorization flow allowing a lead agent to obtain delegation tokens for sub-agents, with shared task context, coordinated scope attenuation, and cross-agent session binding.",
     "https://datatracker.ietf.org/doc/draft-song-oauth-ai-agent-collaborate-authz/",
     "IETF (individual)",
     "Revision -02, Jun/Jul 2026; authors: Yurong Song (Huawei) and colleagues. Addresses the multi-agent coordination gap at the OAuth layer — complements draft-jiang-oauth-intent-admission from the same Huawei group. Distinct from single-hop delegation drafts in explicitly modeling lead-agent-to-sub-agent scope sharing."),

    ("draft-kroehl-agentic-trust-aae — Agent Authorization Envelope (AAE)",
     "Specifies the AAE, a structured authorization container with three mandatory components — MANDATE, CONSTRAINTS, and VALIDITY — providing a machine-evaluable, cryptographically verifiable authorization assertion using W3C DIDs and Verifiable Credentials.",
     "https://datatracker.ietf.org/doc/draft-kroehl-agentic-trust-aae/",
     "IETF (individual)",
     "Revision -00, 21 May 2026; author: Lars Kersten Kroehl (CryptoKRI GmbH). Occupies similar design space to draft-mcguinness-oauth-actor-profile but uses a DID/VC substrate rather than OAuth — provides a useful alternative approach for non-OAuth deployments."),

    ("draft-pidlisnyi-aps — Agent Passport System (APS)",
     "Specifies Ed25519-based agent passports, scoped delegation chains with constraints across seven dimensions (a lattice model where delegation monotonically narrows capabilities), cascade revocation, a three-signature policy chain, signed receipts, and MCP bindings; includes reference implementations in TypeScript and Python.",
     "https://datatracker.ietf.org/doc/draft-pidlisnyi-aps/",
     "IETF (individual)",
     "Revision -01, 14 May 2026; author: Tymofii Pidlisnyi (AEOESS). Among the most comprehensive standalone agent authz specs in the wave — the seven-dimension constraint lattice and MCP binding make it directly implementable against Model Context Protocol deployments alongside draft-serra-mcp-discovery-uri."),

    # ---- Individual drafts: Transport / Infrastructure / Discovery cluster ----
    ("draft-serra-mcp-discovery-uri — The mcp URI Scheme and MCP Server Discovery Mechanism",
     "An IETF individual draft defining an 'mcp://' URI scheme combined with /.well-known/mcp-server (RFC 8615) and a DNS TXT fallback for discovery of Model Context Protocol servers.",
     "https://datatracker.ietf.org/doc/draft-serra-mcp-discovery-uri/",
     "IETF (individual)",
     "Already at revision -04 (March 2026), suggesting active iteration; narrow scope, clean design, and the strongest IANA-registration candidate in the set. OAuth 2.1 for auth is deliberately out of scope."),

    ("draft-sharif-attp — ATTP: Agent Trust Transport Protocol",
     "An IETF individual draft defining trust levels L0–L4 mapped to PSD2 Strong Customer Authentication requirements, with bindings for MCP (MCPS), REST, A2A, gRPC, and GraphQL; supersedes draft-sharif-agent-payment-trust.",
     "https://datatracker.ietf.org/doc/draft-sharif-attp/",
     "IETF (individual)",
     "Revision -00, April 2026; regulatory mapping appendices (PSD2, PCI DSS) are notable — one of the few drafts engaging financial regulation directly. Ambitious scope. Part of the Sharif cluster."),

    ("draft-sharif-agent-transport-protocol — Agent Transport Protocol (ATP): Async Store-and-Forward for AI Agents",
     "An IETF individual draft defining async store-and-forward messaging semantics with cryptographic identity per relay hop, trust scoring at ingress, capability negotiation, and interop with Google A2A, MCP, and FIPA ACL; transport-agnostic over TCP/TLS/QUIC.",
     "https://datatracker.ietf.org/doc/draft-sharif-agent-transport-protocol/",
     "IETF (individual)",
     "Revision -00, 28 March 2026; addresses a real gap — most agent protocols assume synchronous availability, and store-and-forward is appropriate for long-running agentic tasks. Scope is broad and needs narrowing."),

    ("draft-sharif-agent-audit-trail — Agent Audit Trail (AAT): Standard Logging Format for Autonomous AI Systems",
     "An IETF individual draft defining a JSON-based audit record with mandatory fields (agent identity, action classification, outcome tracking, trust level) using RFC 8785 JSON Canonicalization and RFC 9562 UUIDs.",
     "https://datatracker.ietf.org/doc/draft-sharif-agent-audit-trail/",
     "IETF (individual)",
     "Revision -00, 29 March 2026; orthogonal to delegation but important for compliance — standardized audit-log format is a genuine gap. Complements identity frameworks well. A sibling/alternative to the Kuehlewind audit architecture which takes a far more ambitious cross-layer approach."),

    ("draft-aiendpoint-ai-discovery — The AI Discovery Endpoint",
     "An IETF individual draft defining a /.well-known/ai URI suffix for programmatic service capability discovery by AI agents — letting agents discover what an endpoint supports without having to parse human-oriented documentation; requests IANA well-known URI registration.",
     "https://datatracker.ietf.org/doc/draft-aiendpoint-ai-discovery/",
     "IETF (individual)",
     "Revision -00, March 2026; complementary to MCP discovery but broader (not MCP-specific). Needs differentiation from OpenAPI/service-description work to gain traction."),

    ("draft-hood-independent-agtp — Agent Transfer Protocol (AGTP)",
     "An IETF individual draft proposing a dedicated application-layer protocol for AI agent traffic on port 4480 (TCP/TLS and QUIC) with an agtp:// URI scheme, arguing that HTTP is insufficient because agent-generated intent-driven traffic is indistinguishable from human-initiated requests. Defines a Runtime Contract Negotiation Substrate (RCNS) with eighteen methods split into cognitive verbs (QUERY, DISCOVER, DESCRIBE, SUMMARIZE, PLAN, PROPOSE) and mechanics verbs (EXECUTE, DELEGATE, ESCALATE, CONFIRM, SUSPEND, NOTIFY), mandatory agent identity headers, and protocol-level authority scope declaration.",
     "https://datatracker.ietf.org/doc/draft-hood-independent-agtp/",
     "IETF (individual)",
     "Revision -09, 28 Jun 2026 (was -08 May 2026); supersedes draft-hood-independent-atp. Companion specs in the same family: AGTP-API, AGTP-CERT, AGTP-MERCHANT, AGTP-IDENTIFIERS, AGTP-TRUST. Mandatory TLS 1.3+. Ambitious scope — a new transport layer rather than an HTTP profile. The cognitive/mechanics verb split is a distinctive design choice; contrasts with draft-sharif-agent-transport-protocol which layers async store-and-forward on top of existing transports rather than replacing them."),

    ("draft-hood-agtp-ard — ARD Binding for AGTP: Agentic Resource Discovery over the Agent Transfer Protocol",
     "Specifies how Agentic Resource Discovery (ARD) composes with AGTP, defining a catalog entry type for ARD manifests, aligning AGTP's identity model with ARD's trustManifest, and providing an AGTP-native binding for publishing ARD catalogs over AGTP substrate rather than HTTPS.",
     "https://datatracker.ietf.org/doc/draft-hood-agtp-ard/",
     "IETF (individual)",
     "Revision -00, 18 Jun 2026; author: Chris Hood (Nomotic). Extends the AGTP family (draft-hood-independent-agtp in corpus). The trustManifest identity alignment is the piece relevant to the agent identity thread; the rest is transport-layer plumbing for AGTP deployments."),

    ("draft-jernalczyk-intentweb-agent-manifest — IntentWeb AgentManifest",
     "Defines a JSON document that websites publish to describe identity, trusted knowledge, agent-facing capabilities, structured bindings, risk levels, consent requirements, authentication expectations, audit rules, and policies — enabling AI agents to comprehend website functionality and safely perform actions without scraping or fragile UI automation.",
     "https://datatracker.ietf.org/doc/draft-jernalczyk-intentweb-agent-manifest/",
     "IETF (individual)",
     "Revision -00, 6 Jul 2026; author: Mariusz Jernalczyk (IntentWeb). Website-side complement to agent identity drafts; sits in the Discovery & Transport cluster alongside draft-hood-agtp-ard and the DAWN/.well-known discovery thread. Single author, no-affiliation submission — early stage but addresses a real discovery gap."),

    # ---- Individual drafts: Security Requirements / Frameworks cluster ----
    ("draft-ni-a2a-ai-agent-security-requirements — Security Requirements for AI Agents",
     "An IETF individual draft enumerating security requirements for AI agents, covering identity, authorization chaining across domains, and integration points with WIMSE, OAuth identity chaining, and the A2A OAuth profile.",
     "https://datatracker.ietf.org/doc/draft-ni-a2a-ai-agent-security-requirements/",
     "IETF (individual)",
     "Revision -01, February 2026; essential scaffolding for an eventual AI-agent WG. Well-connected to existing IETF work and authored by Ni and Liu, who also lead the WIMSE AI-agent identity draft."),

    ("draft-rosenberg-aiproto-framework — Framework, Use Cases and Requirements for AI Agent Protocols",
     "An IETF individual draft providing a comprehensive AI-agent communications framework that covers user↔agent, agent↔API, and agent↔agent interactions; surveys MCP, A2A, and Agntcy; and sets the stage for IETF standards activity.",
     "https://datatracker.ietf.org/doc/draft-rosenberg-aiproto-framework/",
     "IETF (individual)",
     "Revision -00, October 2025; Rosenberg (Five9) and Jennings (Cisco) are credible long-time IETF contributors. Important landscape document but likely expired (April 2026); watch for a -01 refresh post-IETF 123."),

    ("draft-rosenberg-aiproto-cheq — CHEQ: A Protocol for Confirmation of AI Agent Decisions (HITL)",
     "An IETF individual draft defining an out-of-band Human-in-the-Loop confirmation protocol for agent tool calls, eliminating LLM-hallucination risk for consequential actions and enabling sensitive operations (e.g., banking) without exposing data to the agent.",
     "https://datatracker.ietf.org/doc/draft-rosenberg-aiproto-cheq/",
     "IETF (individual)",
     "Revision -00, October 2025; out-of-band HITL confirmation model is valuable and complements delegation frameworks. Likely expired (April 2026); worth tracking for a -01 refresh."),

    ("draft-sharif-agent-payment-trust — Trust Scoring and Identity Verification for AI Agent Payment Transactions (SUPERSEDED)",
     "An IETF individual draft introducing an Agent Passport — a cryptographic delegation certificate intended to serve as the PSD2 'something you have' factor — and a trust-scoring model for payment transactions mapped to PCI DSS v4.0.1.",
     "https://datatracker.ietf.org/doc/draft-sharif-agent-payment-trust/",
     "IETF (individual, superseded)",
     "SUPERSEDED by draft-sharif-attp-00 (April 2026). Retained for historical context — the Agent Passport concept is interesting but needs formal grounding; the ATTP successor generalizes it across transport bindings."),

    ("draft-stephan-ai-agent-6g — AI Agent Protocols for 6G Systems",
     "An IETF individual draft translating use cases and service requirements from 3GPP TR 22.870 into IETF agent-protocol requirements, covering autonomous vehicles, privacy preservation, and anomaly detection.",
     "https://datatracker.ietf.org/doc/draft-stephan-ai-agent-6g/",
     "IETF (individual)",
     "Revision -02, October 2025; strong operator backing (Orange, Deutsche Telekom, Telefonica, China Mobile, Huawei). Requirements-oriented and important for 3GPP/IETF coordination. May be expired (April 2026); watch for an update."),

    ("draft-jimenez-t2trg-iot-agent — Agentic AI Operation of Constrained RESTful Environments",
     "An IETF individual draft (T2TRG affiliated) describing agentic AI (ReAct, smolagents) operating on CoAP/CoRE constrained-IoT environments, with a reference implementation provided.",
     "https://datatracker.ietf.org/doc/draft-jimenez-t2trg-iot-agent/",
     "IETF (individual)",
     "Revision -00, 14 April 2026; unique in targeting constrained IoT environments. T2TRG affiliation and reference implementation are positive signals. Niche but fills an unaddressed gap."),

    ("draft-li-oauth-delegated-authorization — OAuth 2.0 Delegated Authorization",
     "An IETF individual draft from Huawei authors that extends OAuth 2.0 with a hierarchical 'delegation token' model so a client can mint subordinate, narrowly-scoped access tokens for delegated parties.",
     "https://datatracker.ietf.org/doc/draft-li-oauth-delegated-authorization/",
     "IETF (individual)",
     "Revision -01, March 2026 (Informational); directly addresses over-privileged tokens and is one of the few drafts using the literal term 'delegated authorization'."),

    ("draft-araut-oauth-transaction-tokens-for-agents — Transaction Tokens For Agents",
     "An IETF individual draft from Ashay Raut (Amazon) profiling Transaction Tokens for agent-based workflows by adding 'act' for the agent identity, an 'agentic_ctx' claim carrying agent_type/intent/allowed_actions/environment_constraints, RFC 9396 RAR integration, and an 'actchain' claim for multi-agent delegation lineage.",
     "https://datatracker.ietf.org/doc/draft-araut-oauth-transaction-tokens-for-agents/",
     "IETF (individual)",
     "Revision -01 (May 2026) under the new slug — re-slugged from draft-oauth-transaction-tokens-for-agents-06; the -06 added the actchain delegation lineage, RAR-driven agentic_ctx, and a TTS check that blocks chain-splicing attacks. Conservative extension of the WG Transaction Tokens draft and an easier path to standardization."),

    ("draft-oauth-ai-agents-on-behalf-of-user — On-Behalf-Of User Authorization for AI Agents",
     "An IETF individual draft adding requested_actor and actor_token parameters to the OAuth authorization-code flow so the user gives explicit consent per agent, with the resulting delegation chain documented in the access-token claims.",
     "https://datatracker.ietf.org/doc/draft-oauth-ai-agents-on-behalf-of-user/",
     "IETF (individual)",
     "Revision -02 was submitted August 2025; likely due for -03 refresh. Builds naturally on RFC 6749 + RFC 8693 + RFC 7636 PKCE — one of the most 'IETF-ready' delegation drafts and a strong WG-adoption candidate after the OAuth recharter."),

    ("draft-rosomakho-oauth-txn-challenge — OAuth Transaction Authorization Challenge",
     "An OAuth mechanism enabling protected resources to demand transaction-specific authorization through challenges. When a request involves an agent or automated workflow, the RS returns a challenge to the client; the client presents it to the AS, which validates it, obtains human approval, and issues an access token with RFC 9396 authorization_details describing the approved operation. Positions as complementary to RFC 9470 (Step-Up Authentication) — that spec handles authentication freshness; this one handles authorization specificity.",
     "https://datatracker.ietf.org/doc/draft-rosomakho-oauth-txn-challenge/",
     "IETF (individual)",
     "Revision -00, submitted 25 June 2026. Authors: Yaroslav Rosomakho (Zscaler), Brian Campbell (Ping Identity), Karl McGuinness (Independent), Pieter Kasselman (Defakto). Strong author lineup — Campbell is a prolific OAuth WG contributor, Kasselman co-authored Transaction Tokens and First-Party Apps, McGuinness brings the Mission-Bound OAuth context. The RS-initiated challenge model is distinct from ARAP (which operates at the PDP/PEP layer via AuthZEN) but addresses the same 'denial is not the end' problem — here at the OAuth RS/AS layer rather than the AuthZEN layer."),

    # ---- Individual drafts: Pre-action authorization / permit-before-commit cluster ----
    ("draft-lee-orprg-permit-receipts — Permit Receipts for Permit-Before-Commit Authorization",
     "Defines requirements and an abstract data model for PermitReceipts: a verifier evaluates canonicalized effect requests, action digests, policy epochs, validity intervals, and revocation status before any protected effect is committed at an effect boundary. Covers verifier behavior, failure semantics, and candidate registries.",
     "https://datatracker.ietf.org/doc/draft-lee-orprg-permit-receipts/",
     "IETF (individual)",
     "Revision -00, 4 Jun 2026; author: Yong Bok Lee (Meridian Verity Group). The most complete standalone framework for pre-commit authorization in the current wave; intended to ground IETF discussion on scope and profiles. Seeks guidance on appropriate organizational placement."),

    ("draft-nelson-agent-delegation-receipts — Delegation Receipt Protocol for AI Agent Authorization",
     "Specifies the Delegation Receipt Protocol (DRP) where users sign Authorization Objects (containing scope boundaries, time constraints, instruction hashes, and model state commitments) that are published to an append-only log before the agent runtime gains control. Removes operators as trusted intermediaries by making the user's private key the sole signing authority.",
     "https://datatracker.ietf.org/doc/draft-nelson-agent-delegation-receipts/",
     "IETF (individual)",
     "Revision -10, 13 Jun 2026; author: Ryan Nelson (Authproof). At version 10, the most iterated individual draft in the pre-action authorization space. The commitment-before-execution architecture — sign before the agent runs, not after — is distinctive and directly addresses the prompt-injection attack surface."),

    ("draft-williams-intent-token — The Intent Token: A Cryptographic Authorization Primitive for Autonomous Agents",
     "Specifies a cryptographic authorization primitive that binds an autonomous agent action to a cryptographically signed, human-declared authorization envelope before execution. Addresses the gap in OAuth 2.0 frameworks that govern access to resources but not what agents are permitted to do at the moment of action. Positioned as complementary to OAuth, not a replacement.",
     "https://datatracker.ietf.org/doc/draft-williams-intent-token/",
     "IETF (individual)",
     "Revision -01, 22 Jun 2026; author: Jeffrey Williams (Independent). Overlaps conceptually with draft-kroehl-agentic-trust-aae and draft-nelson-agent-delegation-receipts; worth tracking as a lighter-weight alternative to both."),

    ("draft-pereira-licet-human-intent — LICET: Multi-Modal Physiological Human-Intent Verification for Autonomous AI Agent Authorization",
     "Proposes a biometric HITL protocol fusing ECG, electrodermal activity, and Mahalanobis statistical distance to cryptographically verify genuine human intent — not just credential possession — before authorizing autonomous AI agent actions, with ZK proofs enabling third-party audit without exposing raw biometric data.",
     "https://datatracker.ietf.org/doc/draft-pereira-licet-human-intent/",
     "IETF (individual)",
     "Revision -01, 2 Jul 2026. The most novel and technically niche HITL proposal in the current wave — complements draft-rosenberg-aiproto-cheq (out-of-band HITL confirmation) and draft-sato-soos-hem (kernel-enforced escalation) by grounding human intent in physiological signal rather than credential or UI interaction."),

    # ---- Individual drafts: WIMSE extensions cluster ----
    ("draft-reece-wimse-cross-org-delegation — Cross-Organizational Delegation for Workload and Agent Identity",
     "A problem statement and requirements draft identifying that existing workload and token-based authorization mechanisms were designed for a single trust domain, describing the cross-organizational agent delegation gap, and enumerating requirements for solutions — without proposing specific technical approaches.",
     "https://datatracker.ietf.org/doc/draft-reece-wimse-cross-org-delegation/",
     "IETF (individual)",
     "Revision -00, 25 Jun 2026; author: Morgan Reece (TowerGuardian Consulting). The problem-statement companion to the corpus's operational drafts; likely to feed into WIMSE WG scope expansion and is directly in scope for draft-kuehlewind-audit-architecture's composition graph."),

    ("draft-jiang-wimse-heterogeneous-credential — Heterogeneous Credential Verification for Workload and Agentic Systems",
     "Specifies mechanisms for representing credential sets, identifying credential types, determining appropriate verifiers, normalizing verification results across different formats (workload tokens, OAuth, X.509, attestation), and combining results into a single handling decision.",
     "https://datatracker.ietf.org/doc/draft-jiang-wimse-heterogeneous-credential/",
     "IETF (individual)",
     "Revision -01, 30 Jun 2026 (was -00 Jun 2026); authors: Yuning Jiang, Donghui Wang, Yurong Song, Faye Liu (Huawei). Solves the practical problem of agentic systems receiving credentials in multiple formats from different issuers. Fills a gap left by draft-ni-wimse-ai-agent-identity (corpus) which describes the identity problem but not multi-format credential handling."),

    ("draft-munoz-wimse-authorization-evidence — Signed Authorization-Evidence Records for WIMSE-Authorized AI Agent Actions",
     "Specifies a signed authorization-evidence record (Permit) for WIMSE-authorized AI actions, cryptographically binding to the dispatched request bytes via HTTP Message Signatures, OAuth access tokens, and Shared Signals Framework eventing — without modifying existing standards.",
     "https://datatracker.ietf.org/doc/draft-munoz-wimse-authorization-evidence/",
     "IETF (individual)",
     "Revision -00, 15 May 2026; author: Christian Munoz (Keel API). Profiles WIMSE for AI agent actions and satisfies audit requirements; companion to the SCITT permit profile below. Extends the corpus WIMSE thread (draft-ni-wimse-ai-agent-identity) with a concrete evidence artifact."),

    ("draft-schwenkschuster-wimse-trust-domain-discovery — WIMSE Trust Domain Discovery",
     "Fills a gap in the WIMSE architecture by defining how relying parties obtain cryptographic trust anchors for workload identity credentials; specifies the WIMSE Trust Bundle (a JSON document with freshness metadata) and a well-known HTTPS discovery endpoint resolving trust domain names to their trust bundles, following OpenID Connect Discovery / OAuth 2.0 patterns.",
     "https://datatracker.ietf.org/doc/draft-schwenkschuster-wimse-trust-domain-discovery/",
     "IETF (individual)",
     "Revision -00, 3 Jul 2026; authors: Arndt Schwenkschuster (Defakto Security), Yaroslav Rosomakho (Zscaler). Directly complements draft-schwenkschuster-wimse-credential-exchange (corpus). The kuehlewind-audit-architecture HUB assumes resolvable trust anchors — this provides the discovery mechanism. Rosomakho co-authorship ties it to draft-rosomakho-oauth-txn-challenge."),

    ("draft-winmagic-wimse-condition-bounded-credentials — Condition-Bounded Credentials for Workload and Agent Identity",
     "Proposes hardware-rooted, non-exfiltratable workload credentials whose validity is conditioned on attested runtime posture rather than expiration time — so a credential becomes invalid if the workload's attestation evidence degrades, regardless of the stated expiry.",
     "https://datatracker.ietf.org/doc/draft-winmagic-wimse-condition-bounded-credentials/",
     "IETF (individual)",
     "Revision -01, 6 Jul 2026; author: WinMagic. Extends the WIMSE credential model with posture-conditioned validity — distinct from time-bounded credentials in that security policy changes at the hardware/attestation layer instantly invalidate the credential. Bridges the RATS attestation thread and the WIMSE identity thread in the corpus."),

    # ---- Individual drafts: SCITT / Audit profiles cluster ----
    ("draft-munoz-scitt-permit-profile — A SCITT Profile for Pre-Execution AI Action Authorization Records",
     "SCITT profile for the Pre-Execution Authorization Record (Permit), documenting policy-evaluated decisions before AI agent action dispatch with cryptographic binding proving 'authorized request equals dispatched request.' Composes with adjacent profiles for human-authority binding, post-execution evidence, and content-refusal events via referencing mechanisms.",
     "https://datatracker.ietf.org/doc/draft-munoz-scitt-permit-profile/",
     "IETF (individual)",
     "Revision -00, 15 May 2026; author: Christian Munoz (Keel API). SCITT-anchored analog of the WIMSE evidence record; companion to draft-munoz-wimse-authorization-evidence from the same author. Connects to draft-sharif-agent-audit-trail and draft-kuehlewind-audit-architecture (both in corpus)."),

    ("draft-marques-asqav-compliance-receipts — Compliance Profile of Signed Action Receipts for AI Agents",
     "Establishes a compliance framework for signed action receipts from AI agents, mandating specific fields, cryptographic anchoring, hash-chain linkage, risk/incident classification, cross-agent envelope binding, and enforcement attestation with retention floors tied to EU AI Act, DORA, NIST AI RMF, HIPAA, SEC Rule 17a-4, and CIRCIA.",
     "https://datatracker.ietf.org/doc/draft-marques-asqav-compliance-receipts/",
     "IETF (individual)",
     "Revision -06, 1 Jul 2026 (was -05 May 2026); author: João André Gomes Marques. The most legally grounded receipt profile in the wave — at revision 06, relatively mature. Complements draft-sharif-agent-audit-trail (corpus) with a multi-regulation compliance overlay."),

    ("draft-mih-scitt-agent-action-capsule — An Agent Action Capsule Profile for SCITT",
     "Defines a SCITT statement profile (Agent Action Capsule) documenting agent actions with outcome status (executed, blocked, denied, errored, timed out), evaluated constraints, committed-effect binding, and a human-in-the-loop flag; records a Capsule for every verdict including refusals. COSE Signed Statement format; registrable in SCITT Transparency Services.",
     "https://datatracker.ietf.org/doc/draft-mih-scitt-agent-action-capsule/",
     "IETF (individual)",
     "Revision -02, 6 Jul 2026 (was -01 Jun 2026); author: Steven Mih (Action State Group). -02 added an 'honest human-in-the-loop' flag distinguishing actual HITL from automated policy. Provides auditor-grade evidence that decision gates functioned; the HITL flag directly tracks the corpus theme."),

    ("draft-car-rer-artifact — RER Run Artifact Format: Hash-Chained, Signed Records of AI Inference Execution",
     "Specifies a cryptographically signed JSON record for AI inference runs including a signed envelope declaring permissions and limits, a hash-chained event log covering all model and tool calls, and a runtime signature binding envelope and log to the producing implementation. Enables offline verification by parties uninvolved in the original run.",
     "https://datatracker.ietf.org/doc/draft-car-rer-artifact/",
     "IETF (individual)",
     "Revision -01, 12 Jun 2026; author: Kayla Cardillo (Tech Enrichment). Relevant for post-hoc audit of what permissions an agent claimed during execution; companion to the SCITT profiles but more focused on inference traceability than authorization decisions."),

    ("draft-noa-scitt-ai-agent-receipt — A SCITT Profile for AI-Agent Action Receipts",
     "Defines a minimal SCITT profile: COSE_Sign1 Signed Statements with canonicalized payloads, hash-chained for ordering, registrable in SCITT Transparency Services. Explicitly does NOT assert agent correctness, safety, or real-world outcomes — only tamper-evident, signature-verifiable records of action, principal, policy identity, and verdict.",
     "https://datatracker.ietf.org/doc/draft-noa-scitt-ai-agent-receipt/",
     "IETF (individual)",
     "Revision -00, 23 Jun 2026; author: Tora Toraman (NordenSoft). The minimalist end of the SCITT profile spectrum; useful for corpus completeness alongside the more comprehensive draft-mih-scitt-agent-action-capsule."),

    ("draft-rampalli-scitt-capsule-provenance-binding — Binding Per-Action Authorization and Memory Provenance into Agent Action Capsules",
     "Specifies how to bind three components into AAC records: what was executed ('did'), whether it was authorized ('may'), and the source of the belief that motivated the action ('why-believed'); uses optional payload extensions carrying authorization token references, memory chain roots, and quarantine attestations without modifying AAC core protected-header claims.",
     "https://datatracker.ietf.org/doc/draft-rampalli-scitt-capsule-provenance-binding/",
     "IETF (individual)",
     "Revision -00, 5 Jul 2026; author: Karthik Rampalli (Glyphzero, Inc.). Extends draft-mih-scitt-agent-action-capsule with memory provenance — the 'why-believed' component is novel among SCITT profiles. Includes a defensive publication notice waiving patent claims on the core binding technique."),

    ("draft-nobuo-scitt-composite-evidence-verification — Composite Evidence Verification for SCITT Statement Graphs",
     "Defines a composite verifier function for checking collections of SCITT Signed Statements, receipts, object bindings, and relationship edges under a named verification profile; provides structured reporting on validation status, missing evidence, outdated information, and conflicting claims — targeting auditors needing to evaluate multiple interconnected statements.",
     "https://datatracker.ietf.org/doc/draft-nobuo-scitt-composite-evidence-verification/",
     "IETF (individual)",
     "Revision -00, 7 Jul 2026; author: Nobuo Aoki (SOKENDAI). Operates above individual SCITT profiles (AAC, permits, compliance receipts all in corpus) — the 'how do you verify all of them together' layer that draft-kuehlewind-audit-architecture implies but does not yet specify. Watch for Kühlewind/Birkholz pickup."),

    # ---- Individual drafts: SOOS governance suite (Tom Sato, MyAuberge K.K.) ----
    # Five mutually-referencing drafts forming a coherent agent governance family.
    # Analogous integrator role to draft-kuehlewind-audit-architecture but with a
    # Sovereign Object / WIMSE substrate and explicit EU AI Act Article 12/14 alignment.

    ("draft-sato-soos-mjwt — The Mandate JWT (MJWT) for Agentic AI Systems",
     "Introduces the Mandate JWT — a WIMSE workload credential profile binding agent authority to specific Sovereign Object instances under named human principals, with cryptographically enforced delegation ceilings and a six-dimensional Narrowing Property preventing sub-agents from exceeding root authority.",
     "https://datatracker.ietf.org/doc/draft-sato-soos-mjwt/",
     "IETF (individual)",
     "Revision -01, 10 Jun 2026; author: Tom Sato (MyAuberge K.K.). The authorization token primitive for the SOOS governance protocol family. Directly comparable to draft-mcguinness-oauth-client-instance-assertion but uses a WIMSE/Sovereign Object substrate rather than OAuth."),

    ("draft-sato-soos-mad — Multi-Agent Delegation in Sovereign Object Systems",
     "Specifies three core mechanisms for multi-agent delegation accountability: the Narrowing Property (capability non-amplification), five formally defined object topology patterns for runtime relationships, and cluster coordination primitives for parallel execution with aggregation rules. Version 02 adds revocation-during-execution handling and partial-completion routing to human oversight.",
     "https://datatracker.ietf.org/doc/draft-sato-soos-mad/",
     "IETF (individual)",
     "Revision -02, 10 Jun 2026; author: Tom Sato (MyAuberge K.K.). Claims to address gaps the OpenID Foundation identifies as unsolved in multi-agent authorization. The five topology patterns are a useful vocabulary for describing agent-to-agent relationship structures."),

    ("draft-sato-soos-hem — The Human Escalation Mechanism (HEM) for Agentic AI Systems",
     "Defines a kernel-enforced mechanism placing agent sessions in a pending state when human judgment is required, routing structured escalation requests to designated human decision-makers, and prohibiting all state transitions until a human decision is received. Defines five decision types and a dual-layer architecture (LLM-HEM and SOOS-HEM).",
     "https://datatracker.ietf.org/doc/draft-sato-soos-hem/",
     "IETF (individual)",
     "Revision -05, 30 Jun 2026 (was -04 Jun 2026); author: Tom Sato (MyAuberge K.K.). The most normative HITL protocol in the current wave; explicitly aligns with EU AI Act Article 14 requirements for high-risk AI system oversight. Counterpart to ARAP (AuthZEN WG) at the protocol level vs. the OAuth/AuthZEN layer."),

    ("draft-sato-soos-idp — The Intent Declaration Primitive (IDP) for Agentic AI Systems",
     "Specifies a structured declaration mechanism committing AI agent intent to tamper-evident logs before actions are taken, enabling post-hoc review and EU AI Act Article 12 compliance for high-risk systems.",
     "https://datatracker.ietf.org/doc/draft-sato-soos-idp/",
     "IETF (individual)",
     "Revision -05, 30 Jun 2026 (was -04 Jun 2026); author: Tom Sato (MyAuberge K.K.). The log-before-execute pattern mirrors draft-nelson-agent-delegation-receipts but operates at the intent declaration level rather than the user-authorization level. Pairs with GAR (below) to form a complete pre/post audit trail."),

    ("draft-sato-soos-gar — The Governance Audit Record (GAR) for Agentic AI Systems",
     "Defines GAR, an audit framework with five audit types, a Session Audit Record, and an Audit Alert mechanism; collects, signs, and makes governance events from IDP, HEM, and associated SOOS primitives available for regulatory inspection via an append-only, non-suppressible SCITT-anchored audit stream with Authority Lifecycle Events covering the complete revocation-recovery cycle.",
     "https://datatracker.ietf.org/doc/draft-sato-soos-gar/",
     "IETF (individual)",
     "Revision -03, 28 Jun 2026 (was -02 Jun 10). -03 added OpenTelemetry attribute namespace for governance observability, a GAR Processor spec for converting OTel signals to audit records with integrity verification, four new Authority Lifecycle Event categories (policy conflict scenarios, statutory interpretation changes), and mandatory provenance fields for policy evaluation records. The integrator for the SOOS family, analogous to how draft-kuehlewind-audit-architecture integrates the broader IETF agent draft landscape."),

    # ---- Individual drafts: Security analysis / intent / adjacent protocols cluster ----
    ("draft-jiang-intent-security — Security Considerations and Requirements for Intent-Based Requests in Agentic Systems",
     "Solution-agnostic threat analysis covering tampering, privilege escalation, constraint violations, and intent drift in intent-based agentic systems, with a reference model, attack scenarios, and security requirements covering authentication, admission control, constraint validation, and multi-hop integrity.",
     "https://datatracker.ietf.org/doc/draft-jiang-intent-security/",
     "IETF (individual)",
     "Revision -03, 22 Jun 2026; authors: Yuning Jiang, Lun Li, Yurong Song, Faye Liu (Huawei). The threat-model companion to draft-jiang-oauth-intent-admission from the same group; likely intended to inform the broader IETF agentic AI security discussion."),

    ("draft-somoza-dmsc-atn-agent-trust-negotiation — Agent Trust Negotiation: Capability, Delegation, and Provenance Binding",
     "Specifies the Agent Trust Negotiation (ATN) protocol, which binds four artifacts — capability manifests, delegation chains, provenance attestations, and session receipts — to agent identities via a handshake state machine producing mutually verified, scope-bounded sessions with SCITT-suitable audit records.",
     "https://datatracker.ietf.org/doc/draft-somoza-dmsc-atn-agent-trust-negotiation/",
     "IETF (individual)",
     "Revision -00, 29 May 2026; author: Enrique Somoza (Independent). Operates above discovery mechanisms and positions itself between the discovery layer (DAWN, MCP) and the authorization layer (WIMSE, OAuth). Relevant as an agent-to-agent trust establishment handshake protocol."),

    ("draft-mcgraw-httpapi-agent-budget — The Delegation HTTP Authentication Scheme for Request-Bound Authority",
     "Defines the 'Delegation' HTTP authentication scheme and response semantics for delegated-authority challenges using HTTP status codes and Problem Details, plus a CBOR/COSE proof format. The initial authority profile is 'Budget,' using post-quantum ML-DSA to prove spending or resource-consumption limits.",
     "https://datatracker.ietf.org/doc/draft-mcgraw-httpapi-agent-budget/",
     "IETF (individual)",
     "Revision -02, 15 Jun 2026; author: John Paul McGraw Jr. (TaskHawk Systems). HTTP-layer bounded delegation with post-quantum crypto; distinct from OAuth-layer delegation in that the authority proof travels in the Authorization header. Complements x401 (Industry tab) at the HTTP challenge-response layer."),

    ("draft-vauban-x402-delegation-binding — x402 Delegation Binding for Agentic HTTP Pipelines",
     "Addresses the security gap where x402 V2 payment grants issued to one agent could be misused by another in the same pipeline, introducing three binding mechanisms: a JCS-derived agent identity pseudonym, a nonce for anti-replay, and a depth-limiting field for transitive delegation.",
     "https://datatracker.ietf.org/doc/draft-vauban-x402-delegation-binding/",
     "IETF (individual)",
     "Revision -01, 25 May 2026; author: Vauban Research. Narrowly scoped to payment pipelines but the delegation-binding pattern (pseudonymous identity + anti-replay + depth-limit) is applicable broadly. Validated across five language runtimes with A2A and MCP integration guidance."),

    ("draft-samal-vap — Verifiable Agent Protocol (VAP): Intent-Bound Admission Control and Audit for Agent Tool Invocation",
     "Specifies VAP as a defense-in-depth layer adding declarations of purpose and session budget commitments to tool invocation requests (targeting MCP-style protocols), enabling servers to perform admission control before execution via four messages carried in existing protocol metadata.",
     "https://datatracker.ietf.org/doc/draft-samal-vap/",
     "IETF (individual)",
     "Revision -00, 3 Jun 2026; author: Kruttidipta Samal (Independent). Lightweight and protocol-agnostic; addresses erroneous agent behavior, cost control, and audit trails without replacing existing auth mechanisms. Practical tool-invocation-level authorization layer for MCP deployments."),

    ("draft-khera-aurora — Agent Unification, Runtime, and Operational Responsibility Attestation (AURORA)",
     "A two-layer framework unifying hardware-enclave-backed Runtime Integrity Attestation with scoped, cryptographically bound Authority Delegation; enables agents to prove both operational authority and runtime integrity simultaneously.",
     "https://datatracker.ietf.org/doc/draft-khera-aurora/",
     "IETF (individual)",
     "Revision -00, 5 Jun 2026; author: Ankur Khera. Addresses the gap where delegation tokens don't prove the runtime environment is trustworthy — relevant where delegation must be paired with platform attestation (enterprise/regulated deployments). Bridges the RATS attestation thread and the OAuth delegation thread."),

    ("draft-borthwick-msebenzi-environment-state — Verifiable Intent — environment.* Constraint Family",
     "Defines the environment.* constraint family for agent-authorization mandate vocabularies, covering external conditions at transaction execution time (venue availability, wallet funding) that existing transactional constraints cannot address. Extends agent authorization mandate frameworks with environmental precondition checking.",
     "https://datatracker.ietf.org/doc/draft-borthwick-msebenzi-environment-state/",
     "IETF (individual)",
     "Revision -01, 13 Jun 2026; authors: Douglas Cameron Borthwick (InsumerAPI), Michael Msebenzi (Headless Oracle). Narrow but practically important for financial and commerce agent workflows; connects to draft-mcgraw-httpapi-agent-budget and draft-vauban-x402-delegation-binding."),

    ("draft-chen-oauth-agent-authz-use-cases — Agent Authorization Use Cases and Gap Analysis for OAuth 2.0",
     "A framework-level analysis enumerating agentic authorization use cases — delegated task execution, multi-agent coordination, long-running workflows, capability attenuation — and mapping each against current OAuth 2.0 mechanisms to identify gaps that require new extensions or profiles.",
     "https://datatracker.ietf.org/doc/draft-chen-oauth-agent-authz-use-cases/",
     "IETF (individual)",
     "Revision -01, 5 Jul 2026; useful problem-statement companion to the OAuth recharter's 'Complex Delegation' milestone. Gap analysis directly motivates several other drafts in this corpus — a natural reference document for the OAuth WG scoping discussion."),

    ("draft-agnihotri-oauth-agent-impl-status — Implementation Status of OAuth Identity Chaining and Transaction Tokens",
     "An RFC 7942–compliant implementation status report tracking open-source implementations of draft-ietf-oauth-identity-chaining and draft-ietf-oauth-transaction-tokens against the relevant spec requirements.",
     "https://datatracker.ietf.org/doc/draft-agnihotri-oauth-agent-impl-status/",
     "IETF (individual)",
     "Revision -02, Jun 2026; provides the implementation evidence required to advance both WG drafts toward IESG submission. Informational companion to the two WG drafts — not a protocol spec but tracking their standardization progress."),

    ("draft-hardt-oauth-aauth-protocol — AAuth Protocol",
     "A new clean-sheet authorization protocol from Dick Hardt (original OAuth author) defining proof-of-possession by default, resource-signed challenges, agent identity without pre-registration, deferred 202 responses, and AS-to-AS federation for the agent ecosystem.",
     "https://datatracker.ietf.org/doc/draft-hardt-oauth-aauth-protocol/",
     "IETF (individual)",
     "Revision -09, Jul 4 2026; slug gained 'oauth' infix at this revision (was draft-hardt-aauth-protocol). Now formally adds four resource access modes including agent governance. The design rationale section explicitly explains why AAuth is not GNAP, not OAuth, not DPoP and not mTLS — important read for anyone planning a new agent-auth stack. OUTLIER: zero OAuth dependencies."),

    ("draft-chen-oauth-roadmap — Comprehensive Roadmap for OAuth 2.0 Standards and Drafts",
     "An IETF informational draft that maps the entire OAuth 2.0 ecosystem of RFCs, BCPs, and active drafts into functional layers from core to extensions to industry profiles.",
     "https://www.ietf.org/archive/id/draft-chen-oauth-roadmap-00.html",
     "IETF (individual)",
     "Brand-new -00 draft (May 2026); the single most useful index when starting work in this area."),

    ("Charter: WIMSE — Workload Identity in Multi-System Environments (charter-ietf-wimse-01)",
     "The formal IETF charter for the WIMSE WG covering architecture, JOSE-based WIMSE tokens for service-to-service traffic, local token issuance, and token exchange profiles (likely based on RFC 8693) for cross-trust-domain workload identity.",
     "https://datatracker.ietf.org/doc/charter-ietf-wimse/01/",
     "IETF (WIMSE WG charter)",
     "Approved 18 March 2026 (v01); explicitly liaises with OAuth, SCIM, SCITT, RATS, the OpenID Foundation, and CNCF/SPIFFE — the formal scope statement that anchors all WIMSE draft work."),
]
make_sheet("Active IETF Drafts", COLORS['Drafts'], draft_rows)

# ============================================================
# TAB 3: OpenID Foundation
# ============================================================
oidf_rows = [
    ("OpenID Foundation — Identity Management for Agentic AI (Whitepaper, Oct 2025)",
     "An OpenID Foundation whitepaper arguing that user impersonation by agents should be replaced by delegated authority, requiring explicit 'on-behalf-of' flows where agents prove their delegated scope while remaining identifiable as distinct from the user they represent.",
     "https://openid.net/wp-content/uploads/2025/10/Identity-Management-for-Agentic-AI.pdf",
     "OpenID Foundation",
     "The industry's most-cited problem statement for agent identity; ZeroID and other reference implementations explicitly position themselves against this document."),

    ("OpenID AuthZEN Authorization API 1.0 (Final)",
     "An OpenID Foundation Final Specification that standardizes the request/response between Policy Enforcement Points and Policy Decision Points, often described as 'OIDC for authorization'.",
     "https://openid.net/specs/authorization-api-1_0.html",
     "OpenID Foundation (AuthZEN WG)",
     "Approved as Final 12 January 2026 by 81–1–25 vote; complements OAuth's delegation by externalizing the runtime allow/deny decision."),

    ("OpenID AuthZEN Profile for MCP Tool Authorization",
     "An OpenID AuthZEN profile that maps MCP tool-invocation authorization onto the AuthZEN PDP/PEP API so MCP servers can externalize per-tool decisions.",
     "https://github.com/openid/authzen",
     "OpenID Foundation (AuthZEN WG)",
     "Among the first standards-body responses to the agent/MCP authorization problem; lives in the AuthZEN GitHub alongside the core spec."),

    ("OpenID AuthZEN Access Request and Approval Profile (ARAP) — Draft 1",
     "An AuthZEN WG draft (adopted May 2026, Draft 1 published 3 June 2026, sole author Karl McGuinness) that lets a PDP return decision:false with a structured 'access_request' context object pointing at an Access Request Endpoint. The PEP submits the request, gets an opaque Task Handle, polls or receives callbacks, and re-evaluates against the PDP after the workflow completes — denial remains denial; the PDP stays authoritative.",
     "https://openid.github.io/authzen/authzen-access-request-approval-profile-1_0.html",
     "OpenID Foundation (AuthZEN WG)",
     "The standardization of 'deny, but escalate' as a runtime primitive. Concepts: evaluation_id binding, JWS binding_token, approval.state (also JWS), bulk items[] submissions, catalog references for entitlement enumeration, multi-step approval progress, cancellation, idempotency. Single completion mode in the base profile (reevaluate); other modes deferred to follow-on profiles. AI agents are a primary use case in the worked examples (the 'Agent Tool Discovery' end-to-end appendix). Composes with McGuinness's Actor Profile draft for the act-chain shape. Explicitly positioned against CIBA: ARAP solves authorization escalation, CIBA solves authentication freshness — different problems."),

    ("OpenID Continuous Access Evaluation Profile (CAEP) 1.0",
     "An OpenID Foundation specification that defines event types over the Shared Signals Framework so transmitters can asynchronously notify receivers of session, credential, or device-posture changes.",
     "https://openid.net/specs/openid-caep-1_0-final.html",
     "OpenID Foundation (Shared Signals WG)",
     "Pairs with AuthZEN to give continuous (not just point-in-time) authorization — the async layer of the 'continuous authorization loop'."),

    ("OpenID Shared Signals Framework (SSF) 1.0",
     "An OpenID Foundation specification that defines a generic transport for Security Event Tokens (RFC 8417) between cooperating identity providers and relying parties.",
     "https://openid.net/wg/sharedsignals/specifications/",
     "OpenID Foundation (Shared Signals WG)",
     "Underlies both CAEP (session/access changes) and RISC (account-compromise signals); 9-vendor interop demonstrated at Gartner IAM London 2025."),

    ("CAEP Interoperability Profile 1.0",
     "An OpenID profile that pins down concrete bindings, OAuth 2.0 usage, and minimal event-type support so independent SSF/CAEP implementations can actually interoperate.",
     "https://openid.github.io/sharedsignals/openid-caep-interoperability-profile-1_0.html",
     "OpenID Foundation (Shared Signals WG)",
     "Critical because the CAEP base spec leaves enough optionality that vendors weren't truly interoperable until this profile."),

    ("FAPI 2.0 Security Profile (Final)",
     "The OpenID Foundation Final Specification that defines a high-security OAuth 2.0 profile requiring sender-constrained tokens (DPoP/mTLS), PAR, and formally analyzed properties.",
     "https://openid.net/specs/fapi-security-profile-2_0-final.html",
     "OpenID Foundation (FAPI WG)",
     "Approved Final Feb 2025; formally verified by University of Stuttgart and now the de-facto baseline for open banking/open data globally."),

    ("FAPI 2.0 Message Signing",
     "An OpenID FAPI extension that adds non-repudiation by requiring signatures over authorization requests, responses, and resource-server messages on top of the FAPI 2.0 baseline.",
     "https://openid.net/specs/fapi-2_0-message-signing-ID1.html",
     "OpenID Foundation (FAPI WG)",
     "Final conformance tests launched August 2025; required where regulators or schemes demand non-repudiation."),

    ("OpenID FAPI Grant Management for OAuth 2.0",
     "An OpenID FAPI specification that standardizes how clients create, query, update, and revoke long-lived consent grants, born from PSD2 and Australian CDR experience.",
     "https://openid.net/wg/fapi/specifications/",
     "OpenID Foundation (FAPI WG)",
     "Solves the recurring 'where do I see and revoke what I've delegated?' UX problem at scale."),

    ("Health Relationship Trust Profile for UMA 2.0 (HEART UMA2)",
     "An OpenID HEART profile of UMA 2.0 that tightens cryptographic and consent-management requirements for healthcare/HIPAA-style multi-party API scenarios.",
     "https://openid.net/specs/openid-heart-uma2-1_0.html",
     "OpenID Foundation (HEART WG)",
     "Reference for any patient-mediated data-sharing implementation, including 21st Century Cures Act deployments."),
]
make_sheet("OpenID Foundation", COLORS['OpenID'], oidf_rows)

# ============================================================
# TAB 4: Other standards bodies
# ============================================================
other_rows = [
    ("User-Managed Access (UMA) 2.0 Grant for OAuth 2.0 Authorization",
     "A Kantara Initiative recommendation defining an OAuth 2.0 extension grant for asynchronous, party-to-party delegation where the resource owner pre-configures policy on the AS.",
     "https://docs.kantarainitiative.org/uma/wg/rec-oauth-uma-grant-2.0.html",
     "Kantara Initiative",
     "The original 'Alice-to-Bob' delegated-authorization standard; broadly supported by Keycloak, ForgeRock/Ping, WSO2."),

    ("W3C Verifiable Credentials Data Model v2.0 (Recommendation)",
     "The W3C Recommendation defining a tamper-evident, cryptographically-signed claim structure with an issuer/holder/verifier model and support for selective disclosure.",
     "https://www.w3.org/TR/vc-data-model-2.0/",
     "W3C (VC WG)",
     "Reached Recommendation status 15 May 2025; underlies EUDI Wallet, mDL, and most agent-attestation proposals."),

    ("W3C Verifiable Credentials Data Model v2.1 (First Public Working Draft)",
     "The W3C VC Working Group's First Public Working Draft of the next data-model revision, refining v2.0 alignment with JOSE/COSE and Data Integrity proofs.",
     "https://www.w3.org/news/2026/first-public-working-draft-verifiable-credentials-data-model-v2-1/",
     "W3C (VC WG)",
     "Published April 2026; planned EU recognition target April 2027 per VC WG charter."),

    ("W3C Verifiable Credentials Working Group Charter (2026)",
     "The current W3C charter for the VC WG that schedules upcoming Recommendations on Confidence Method, Rendering Methods, and an HTTP API for issuance and presentation.",
     "https://w3c.github.io/vc-charter-2026/",
     "W3C",
     "Confidence Method and Rendering Methods exclusion periods ended 29 March 2026 — these are the v2.1 sibling specs to watch."),

    ("NIST AI Agent Standards Initiative — Center for AI Standards and Innovation",
     "A NIST initiative within the Center for AI Standards and Innovation aimed at accelerating adoption of software and AI-agent identity and authorization standards.",
     "https://www.nist.gov/ai",
     "NIST (US government)",
     "February 2026 NCCoE concept paper on 'Accelerating the Adoption of Software and AI Agent Identity and Authorization' is the key entry point."),

    ("EU AI Act broader enforcement (high-risk AI systems audit trails, August 2026)",
     "EU regulation requiring auditable trails for high-risk AI systems, which in practice mandates the kind of cryptographic delegation chains being standardized in IETF agent drafts.",
     "https://artificialintelligenceact.eu/",
     "European Union (regulation)",
     "Compliance deadline shapes urgency for delegation-chain standardization; many vendors target Aug 2026 readiness."),
]
make_sheet("Other Standards & Govt", COLORS['Other'], other_rows)

# ============================================================
# TAB 5: Academic
# ============================================================
academic_rows = [
    ("Authenticated Delegation and Authorized AI Agents (South et al., arXiv:2501.09674)",
     "An academic paper from MIT, Anthropic, and collaborators proposing an OAuth/OIDC-compatible framework for authenticated, scoped, auditable delegation of authority from humans to AI agents.",
     "https://arxiv.org/abs/2501.09674",
     "Academic (arXiv preprint)",
     "Highly cited 2025 paper that re-framed agent authorization as 'delegation with chains of accountability' and influenced subsequent IETF drafts."),

    ("Delegated Authorization for Agents Constrained to Semantic Task-to-Scope Matching (arXiv:2510.26702)",
     "An academic paper introducing a model in which the AS semantically inspects an agent's intended task and grants the minimal scope set, plus the ASTRA benchmark dataset.",
     "https://arxiv.org/abs/2510.26702",
     "Academic (arXiv preprint)",
     "Published Oct 2025; one of the first works giving researchers a public benchmark for semantically-grounded delegated authorization."),

    ("Establishing Workload Identity for Zero Trust CI/CD (Avirneni, arXiv:2504.14760)",
     "An academic paper describing the migration from static CI/CD secrets through OIDC federation to runtime SPIFFE-issued workload identities for non-human actors.",
     "https://arxiv.org/abs/2504.14760",
     "Academic (arXiv preprint)",
     "Concrete enterprise-grade reference implementation tying SPIFFE/SPIRE to OIDC federation in GitHub Actions/AWS/GCP/Azure."),

    ("Identity Control Plane: The Unifying Layer for Zero Trust Infrastructure (arXiv:2504.17759)",
     "An academic position paper proposing an Identity Control Plane that unifies SPIFFE workload identity, OIDC/SAML human identity, and broker-issued transaction tokens under one policy plane.",
     "https://arxiv.org/abs/2504.17759",
     "Academic (arXiv preprint)",
     "Useful conceptual framing because most enterprises already have all three identity types but no unified plane."),

    ("Formal Security Analysis of the OpenID Financial-grade API 2.0 (Hosseyni, Kuesters, Würtele, IEEE CSF 2024)",
     "An academic paper providing the formal-methods proof of FAPI 2.0's security properties under the FAPI 2.0 attacker model, performed by the University of Stuttgart team.",
     "https://doi.ieeecomputersociety.org/10.1109/CSF61375.2024.00002",
     "IEEE CSF (Academic conference)",
     "The proof underpinning FAPI 2.0's claim of being formally verified; the same authors now drive the OAuth Security Topics Update draft."),
]
make_sheet("Academic Papers", COLORS['Academic'], academic_rows)

# ============================================================
# TAB 6: Industry & Implementations — with ZeroID and auth.md
# ============================================================
industry_rows = [
    # ---- Reference implementations of the standards stack ----
    ("OVID-ME — Cedar Policy Evaluation for OVID Agent Mandates",
     "A reference implementation that enforces attenuated multi-hop agent-delegation mandates by evaluating Cedar policies embedded as RFC 9396 authorization_details inside signed OVID JWTs.",
     "https://github.com/clawdreyhepburn/ovid-me",
     "Open-source (Apache-2.0)",
     "Companion to the OVID identity package; provides AuthZEN-compliant PDP, dry-run/shadow/enforce modes, and SMT-based subset proof at issuance time. Pairs with Carapace as a two-layer enforcement stack."),

    ("OVID — Cryptographic Agent Identity (npm @clawdreyhepburn/ovid)",
     "An Ed25519 JWT package giving each agent a signed cryptographic identity with walkable delegation chains back to the human, designed to compose with OVID-ME at evaluation time.",
     "https://www.npmjs.com/package/@clawdreyhepburn/ovid",
     "Open-source (Apache-2.0)",
     "Implements the SPIFFE-style 'spawner is the attestor' model for agents; conceptual sibling of WIMSE workload tokens for the agent-delegation use case."),

    ("ZeroID — Autonomous Agent Identity Management System (AAIMS)",
     "A Go-based open-source identity service from Highflame that issues short-lived agent credentials over OAuth 2.1 with WIMSE/SPIFFE URIs, RFC 8693 delegation with automatic scope attenuation, configurable max delegation depth, and CAEP/SSF cascading real-time revocation.",
     "https://github.com/highflame-ai/zeroid",
     "Open-source (Apache-2.0)",
     "v1.1.11 released March 2026; one of the most complete production-grade reference implementations of the agent-identity stack — explicitly cites the OpenID Foundation's Oct 2025 Agentic AI whitepaper as its design north-star."),

    ("auth.md — Agentic Registration Protocol (WorkOS)",
     "A reference implementation of a 'robots.txt for agent authentication': an AUTH.md skill manifest at a service's domain that tells agents how to register via three discovery-driven flows — identity assertion (ID-JAG), verified-email assertion, or anonymous OTP claim.",
     "https://github.com/workos/auth.md",
     "Open-source (MIT, WorkOS)",
     "Composes directly with draft-ietf-oauth-identity-assertion-authz-grant; the discovery and onboarding piece that the IETF stack doesn't define, exposed via /.well-known/oauth-authorization-server with an agent_auth block."),

    ("ATP — Agent Trust Protocol Core (atp-sdk)",
     "A TypeScript/Node SDK plus multi-service Docker stack providing quantum-safe (hybrid Ed25519 + NIST ML-DSA) agent identity over a did:atp DID method, with continuous trust scoring, zero-knowledge trust-level proofs, and first-class adapters for LangChain, Motleycrew, MCP, Swarm, ADK, and A2A.",
     "https://github.com/agent-trust-protocol/atp-core",
     "Open-source (Apache-2.0, Larry Lewis)",
     "Specification implicit in implementation — no separate spec doc. Differentiators: PQ hybrid signatures via FIPS 204 ML-DSA, ZK proofs for trust-level predicates (prove trust ≥ 0.7 without revealing the score), multi-framework adapters as a first-class concern. Acronym near-collides with draft-sharif-attp (Agent Trust Transport Protocol) — same problem space, distinct designs: discrete L0-L4 trust levels there, continuous 0.0-1.0 trust scoring here. Early-stage (1 star, 2 contributors, one of whom is Claude); README ships a real Context7 API key, a 'ship-fast' signal worth noting."),

    ("ADCS — Agent Delegation Chain Standard v1.0.0 (kahalewai)",
     "A vendor-neutral open specification for managing permissions across multi-agent AI workflows using cryptographic verification at each delegation hop. Core principle: scope can only narrow, never expand (monotonic restriction). Three cumulative conformance levels (core data model → JWT tokens → production infrastructure), a four-phase verification algorithm (structural, cryptographic, temporal, monotonic), seven standardized constraint types, and DID-based signing. Protocol bindings for MCP, A2A, ACP, and HTTP.",
     "https://github.com/kahalewai/adcs/blob/main/spec/ADCS-Standard-v1.0.0.md",
     "Open-source spec (Apache-2.0)",
     "Version 1.0.0 Public Draft for Community Review, dated 2026-04-11. Spec-only repository — no separate reference implementation. Advanced features include chain compaction, CAEP continuous revocation, multi-party quorum, and permission escalation. The monotonic-restriction guarantee is the same invariant as OVID-ME's Cedar-based subset proof and ZeroID's automatic scope attenuation — three independent approaches to the same core security property."),

    ("x401: HTTP Proof Requirement Protocol (Proof / Circle)",
     "A community specification defining an HTTP-native wrapper for credential-based proof requirements, using three dedicated headers — PROOF-REQUIRED, PROOF-PRESENTATION, and PROOF-RESPONSE — to gate access to protected resources. The verifier encodes a W3C Digital Credentials API / OpenID4VP / DCQL request in PROOF-REQUIRED; the agent acquires and presents a verifiable presentation; an optional fourth leg exchanges the VP for a reusable OAuth 2.0 token. Designed to be stateless at the verifier, transport-agnostic, and composable with existing credential protocols rather than replacing them.",
     "https://x401.proof.com/spec/latest",
     "Community spec (Proof / Circle)",
     "Version 0.2.0, Draft. Editors: Daniel Buchner (Proof), Bhushit Agarwal (Circle); contributors from Google, Okta, OpenAI, MATTR. Explicitly does not replace OpenID4VP or the W3C Digital Credentials API — fills the HTTP-layer gap between those protocols and application-level authorization. Analogous positioning to how RFC 9728 (OAuth Protected Resource Metadata) provides the HTTP-layer discovery wrapper for OAuth; x401 provides the HTTP-layer challenge wrapper for VC presentations. Agent authentication and binding are composable add-ons, not mandatory."),

    # ---- LinkedIn-style and analyst articles ----
    ("AI Agents and the Multi-Hop Delegation Problem (WorkOS)",
     "A WorkOS engineering blog post that catalogs the open IETF drafts and RFCs addressing multi-hop AI-agent delegation including identity chaining, attenuating tokens, and actor chains.",
     "https://workos.com/blog/oauth-multi-hop-delegation-ai-agents",
     "Industry blog (WorkOS)",
     "Particularly useful index of currently-active drafts (April 2026) on the agent-delegation problem with NIST and EU AI Act compliance context."),

    # ---- McGuinness "Mission-Bound OAuth" blog series (May–Jun 2026) ----
    # Karl McGuinness (former Okta SVP & Chief Product Architect) — a four-post architectural argument
    # for treating the user-approved task itself as a first-class OAuth object (the "Mission"). Companion
    # to draft-mcguinness-oauth-actor-profile, draft-mcguinness-oauth-client-instance-assertion, and the
    # AuthZEN ARAP profile listed in the OpenID Foundation tab. Listed in publication order.

    ("Mission-Bound OAuth MVP (McGuinness, 22 May 2026)",
     "The protocol-level proposal. Five wire additions on top of existing OAuth: a 'mission_intent' RAR envelope (purpose, mission_expiry, context), a generic 'resource_access' RAR type, a durable Mission record at the Authorization Server, an opaque 'mission' claim (id + origin) on access tokens, and a Mission-state enforcement gate on refresh / exchange / introspection / assertion validation. Seven-state Mission lifecycle (pending_approval, active, suspended, revoked, expired, completed, rejected). proposal_hash (SHA-256 over JCS-canonical authorization_details) and consent_rendering_hash anchor what was approved versus what the user saw.",
     "https://notes.karlmcguinness.com/notes/mission-bound-oauth-mvp/",
     "Architect blog (Karl McGuinness)",
     "76-min read; the substrate post for the whole series. Target I-D name: draft-mcguinness-oauth-mission-bound-minimum-profile. Conformance Ladder L0–L5 (L0 baseline OAuth → L5 verifiable governance with portable receipts). Cross-AS handoff uses ID-JAG for user-rooted flows or the Fletcher Transaction Token Chaining Profile for Txn-Token-rooted flows. Mission Expansion creates a successor Mission with 'mission.supersedes' rather than mutating in place. Three Resource Server tiers (RS-A OAuth-only → RS-D Mission-state aware via introspection or SSF/CAEP events). Architectural challenges acknowledged honestly: state-sync at scale, unknown-constraint brittleness, lethal-trifecta boundary."),

    ("The Mission is the Missing OAuth Abstraction (McGuinness, 1 Jun 2026)",
     "The architectural frame: OAuth has no first-class object for 'the task the user approved' — only tokens, scopes, prompts, and logs that are downstream projections of it. The Mission is that durable, AS-stored, user-approved authority record. Argues this is what closes the gap that five prior bodies of his work (Power of Attorney, Mission Shaping, Open-World OAuth, Sessions Are Not Missions, the Mission-Bound architecture series) have circled from different angles. Two layers, one object: issuance-bound authority (MVP) plus runtime-enforced authority (the IBAC profile).",
     "https://notes.karlmcguinness.com/notes/the-mission-is-the-missing-oauth-abstraction/",
     "Architect blog (Karl McGuinness)",
     "Short (9-min) framing post that ties the whole programme together. The argument for why IBAC becomes practical when intent is compiled from an AS-validated Mission at consent time rather than inferred post-hoc from agent behaviour (where it's adversarial-input territory and the PDP has no user to ask). Best entry point for someone new to the series."),

    ("Mission-Bound OAuth Runtime Enforcement Profile (McGuinness, 1 Jun 2026)",
     "The IBAC layer layered on the MVP. Core (required for compliance): Intent-to-Policy Compilation (AS deterministically compiles approved authorization_details to an evaluable artifact at activation, stores policy_version on the Mission), Resource-Side Enforcement Contract (RS-B minimum, PDP evaluates every consequential request), Standard Subset Semantics per RAR Type with strict-refuse on unknown constraints (stricter than MVP's 'preserve or refuse'), Mission Introspection Profile (extended response with act chain, tenant, subject, policy_version), Runtime Denial and Escalation via ARAP (MUST), Local-Action Boundary requiring AuthZEN Access Evaluation for non-OAuth actions, Parameter Binding / TOCTOU Protection (parameter_digest bound to the permit), Decision Evidence Records (per-decision audit record bound to mission.id, proposal_hash, policy_version, decision, constraint clauses, act chain). Six Optional Modules: Tool Binding Profile, Decision Receipt Profile (W3C VC 2.0), Actor Provenance Profile, Purpose Registry Profile, Attestation Profile (RATS PTV + WIMSE), Policy Projection Profile (Cedar carriage).",
     "https://notes.karlmcguinness.com/notes/mission-bound-oauth-runtime-enforcement-profile/",
     "Architect blog (Karl McGuinness)",
     "37-min read; target I-D name: draft-mcguinness-oauth-mission-bound-runtime-enforcement-profile. Six-class action classification (non-consequential → consequential read → consequential write → irreversible → external commitment → privileged administration) determines PDP-gate requirement and parameter binding. Four PDP deployment modes (AS-hosted, RS-hosted, tenant governance, federated). Goal pair: 'execution continuity' (every in-bounds action succeeds; every out-of-bounds becomes governed Mission Expansion) plus 'proof of authority' (per-decision cryptographic receipts). Acknowledges PDP latency overhead and tool-manifest fracturing as real challenges."),

    ("Authorization Denied Is No Longer Enough (McGuinness, 2 Jun 2026)",
     "The framing post for ARAP. In closed-world authorization, 'decision:false' was the end of the interaction. In open-world agentic systems with runtime discovery, sub-agent delegation, and evolving missions, denial is increasingly the beginning of a governance escalation — and the missing protocol primitive is a 'requestable denial': a deny that names where to ask and binds the request to the exact evaluation it remediates. Why CIBA isn't the answer (CIBA solves authentication freshness; this is about governance state). Why approvals aren't authority (reevaluation against current state, not standing entitlement).",
     "https://notes.karlmcguinness.com/notes/authorization-denied-is-no-longer-enough/",
     "Architect blog (Karl McGuinness)",
     "Confirms that the AuthZEN ARAP profile was adopted as a working group draft in May 2026 — material maturity signal. Useful as the 'why' read alongside the ARAP spec itself. Also positions the work against AARM (Autonomous Action Runtime Management, aarm.dev/spec) which intercepts every agent action and resolves to allow/deny/modify/step-up/defer — ARAP standardizes the deny+escalate boundary at the AuthZEN layer."),

    ("Re-Subjecting Is a Mint, Not an Attenuation (McGuinness, 8 Jun 2026)",
     "Argues that crossing subject namespaces — mapping a user's identity from one application's identifier to another's — is a minting operation, not attenuation. Attenuation can narrow authority already represented by an existing artifact; it cannot authoritatively create a target-local identity binding the original issuer never supplied. Only a trusted IdP or broker can perform that translation. Distinguishes two topologies: caller-pushed (intermediate app returns to the IdP for a new assertion) and resource-pulled (destination resolves user identity through a broker). Separates workload identity (which service is calling) from user context (which person delegated the work) as requiring distinct claims.",
     "https://notes.karlmcguinness.com/notes/re-subjecting-is-a-mint-not-an-attenuation/",
     "Architect blog (Karl McGuinness)",
     "Published 8 Jun 2026; a standalone conceptual post outside the four-part Mission-Bound series but thematically continuous with it. Directly relevant to cross-AS delegation flows in the MVP post (ID-JAG for user-rooted re-subjecting) and to the token-exchange-target-service-discovery I-D. Privacy note: every intermediary-visible artifact should minimize identifiers that enable unauthorized cross-context linking."),

    ("Solving the Identity Crisis for AI Agents (Uber Engineering)",
     "A production engineering post from Uber describing their agent identity architecture: an Agent Registry (workload-to-agent mapping), a SPIRE-backed STS that issues short-lived JWTs with embedded actor chains at P99 below 40ms, an MCP Gateway as policy enforcement point, and an AI Agent Mesh for agent-to-agent communication. A standardized A2A client automates token exchange and chain propagation across agent hops.",
     "https://www.uber.com/us/en/blog/solving-the-agent-identity-crisis/",
     "Industry blog (Uber Engineering)",
     "Published 21 May 2026; six-author post (Mathew, Borole, Huang, Burykin, Goel, Walsh) from Uber's platform engineering team. Architecture aligns with WIMSE workload identity (SPIRE as the credential foundation), RFC 8693 actor-chain propagation, and single-hop short-lived JWTs with audience-scoped claims. One of the few public disclosures of a production-grade agent identity system at hyperscaler scale — a real-world reference point for the corpus's delegation-chain design patterns."),

    ("Agents Are Not Just Workloads (Patrick Parker, LinkedIn)",
     "Argues that classifying AI agents as 'workloads' is a category error that corrupts identity architecture: workloads are a what-runs-where concept, agents are a who-acts-on-whose-authority concept. Identifies five breakdown areas where workload-identity patterns fail for agents: intent captured as static token claims, delegation gaps (no purpose binding or revocation paths), overlooked tool-catalog authorization surface, bearer-credential custody ambiguity under prompt injection, and mutable logs insufficient as tamper-evident authorization evidence.",
     "https://www.linkedin.com/pulse/agents-just-workloads-patrick-parker-0qxte/",
     "Industry blog (LinkedIn)",
     "Published 9 Jun 2026. Proposes signature-based receipts binding relationships, authority, tasks, and bounds; runtime validation of generated task intent; and AuthZEN gateway-based enforcement. A principled counterpoint to the WIMSE-as-sufficient-for-agents framing — pairs well with the Uber Engineering post (which uses SPIRE/WIMSE as the credential foundation but adds actor chains and MCP gateway enforcement on top)."),

    ("Creating a Relationship Binding (Tom Jones)",
     "A community whitepaper proposing a signed 'Consent to Create Binding' JWT/JWS message that establishes a subject-bound relationship identifier (UUID or DID) between a user and a set of agents or service providers. The message carries 16 mandatory/optional fields: issuer, subject, subject role, context (trust framework), permissions, device statement, identity proof, purpose of use, key material, and signature. Includes a lifecycle termination model (session, cookie, transaction, relationship, and legal-based retention scopes) and a companion 'Consent Receipt with Binding' response. A second section argues that relationship-based governance — binding actors through mutual obligations, escalation paths, and accountability chains — is more resilient for AI agent systems than representational approaches (policies, classifications, static maps).",
     "https://docs.google.com/document/d/1CwBDRbw147YlNld-UOlJ4lGvmpDIOCSI/edit#heading=h.kp911j18z01y",
     "Community whitepaper",
     "No publication date in the document; primary use case is healthcare (NIST IAL2/AAL2). Draws on Kantara's Distribute Assurance Specification for identity assurance levels and references OpenID Connect id_token as the closest existing message shape. The Consent to Create Binding message adds purpose of use, device statement, and lifecycle termination semantics that id_token lacks. Companion doc 'Digital Contracts Made Legally Valid' (separate Google Doc) covers giving the binding legal status. Relationship-governance argument is thematically aligned with Parker's 'Agents Are Not Just Workloads' — both argue that accountability chains and mutual obligations are the right governance primitive for agentic systems."),

    ("Intent Agent Native Authorization for Agentic Profiles (IANA-AP) (Martin Besozzi, Jul 2026)",
     "Proposes a four-phase runtime authorization lifecycle for AI agents: discovery (agent capabilities + API authorization mappings), intent computation (agent generates planned operations), review/approval (user provides explicit phishing-resistant consent via WebAuthn/FIDO2 passkeys before execution), and enforcement (every runtime invocation validated against the approved authorization artifact). Composes RFC 9396 RAR, OAuth FiPA, OpenID AuthZEN, MCP, SPIFFE JWT-SVIDs, and CEL policy mappings into a single coherent framework.",
     "https://embesozzi.github.io/posts/martin-besozzi/intent-agent-native-authorization-agentic-profiles/",
     "Architect blog (Martin Besozzi)",
     "Updated July 2026; includes open specification and reference implementation on GitHub. The SARC (Subject, Action, Resource, Context) model and 'x-authz-mapping' Agentic Profile extension are novel additions. The four-phase pre-execution approval requirement directly parallels draft-nelson-agent-delegation-receipts and draft-sato-soos-idp — but grounds it in existing OAuth standards rather than new protocol primitives. One of the few practitioner-level posts that explicitly composes AuthZEN, RAR, and FiPA into a working agent authorization flow."),

    ("AI Agent Authentication Gets the Hard Part Right. Authorization Is Still Your Problem. (Rock Cyber Musings)",
     "An analyst article arguing the new IETF agent-auth draft solves authentication via SPIFFE+WIMSE+OAuth but leaves authorization and policy enforcement as an unsolved open problem.",
     "https://www.rockcybermusings.com/p/i-agent-authentication-authorization-gap",
     "Industry analysis blog",
     "Maps the IETF draft against MCP and the Colorado AI Act 'reasonable care' standard; pragmatic implementer perspective."),

    ("Agent Authentication & Delegated Access (Zylos Research, April 2026)",
     "A research-style industry article surveying OAuth flows, scoped tokens, and identity patterns specifically for AI agents, covering OAuth 2.1 baselines and chain-splicing risks.",
     "https://zylos.ai/research/2026-04-11-agent-authentication-delegated-access-oauth-scoped-tokens",
     "Industry research blog",
     "Documents the 'delegation chain splicing' attack against RFC 8693 actor-token chains formally raised on the OAuth WG list in early 2026."),

    ("AuthZEN + Shared Signals Framework Series (Andrew Doering)",
     "A multi-part technical blog walking through how AuthZEN (synchronous PDP/PEP) and SSF/CAEP (asynchronous events) combine into a continuous-authorization loop in real Microsoft Entra/M365 deployments.",
     "https://andrewdoering.org/blog/2026/authzen-shared-signals-framework-part-1-fundamentals/",
     "Industry blog",
     "Documents the 2026 asymmetry where Microsoft Entra is a CAEP transmitter for closed-loop CAE but not an external SSF receiver."),

    ("Just-in-Time Authorization with OpenID SSE and CAEP (The New Stack, Tulshibagwale)",
     "An article from CAEP's original inventor explaining how SSE+CAEP enable just-in-time authorization decisions instead of long-lived session-based access.",
     "https://thenewstack.io/just-in-time-authorization-with-openid-sse-and-caep/",
     "Industry article (The New Stack)",
     "Background piece by Atul Tulshibagwale (SGNL CTO, CAEP inventor) explaining the architectural intent."),

    ("How AuthZEN, Shared Signals & CAEP Complement Each Other (OpenID Foundation)",
     "An OpenID Foundation explainer arguing that AuthZEN and SSF/CAEP are complementary — sync access decisions vs async session updates — not competing standards.",
     "https://openid.net/how-authzen-and-shared-signals-caep-complement-each-other/",
     "OpenID Foundation",
     "Useful when explaining the layering to stakeholders confused by the OpenID alphabet soup."),

    ("OAuth 2.1 Features You Can't Ignore in 2026 (Gutierrez, Medium)",
     "A practitioner article summarizing why OAuth 2.1 — mandated PKCE, exact redirect matching, sender-constrained tokens — is the 2026 minimum bar for delegated authorization.",
     "https://rgutierrez2004.medium.com/oauth-2-1-features-you-cant-ignore-in-2026-a15f852cb723",
     "Industry blog (Medium)",
     "Author is Cyber Intelligence Lead — IAM/AI at Oracle; concise piece suitable for executive briefings."),

    ("RFC 9396: OAuth 2.0 Rich Authorization Requests (CIAM Weekly)",
     "A practitioner deep-dive on RAR explaining why scopes alone don't carry transaction-level intent and how authorization_details is being adopted in payments and verifiable credentials.",
     "https://ciamweekly.substack.com/p/rfc-9396-oauth-20-rich-authorization",
     "Industry newsletter",
     "Captures the current (March 2026) state of RAR adoption, including CAMARA telco APIs and DPV-purpose binding proposals."),

    ("Technical Deconstruction of MCP Authorization (kane.mx, Nov 2025)",
     "A long-form technical article showing that the Model Context Protocol's authorization spec is, in practice, an OAuth 2.1 profile combined with RFC 9728 protected-resource metadata and RFC 7591 dynamic registration.",
     "https://kane.mx/posts/2025/mcp-authorization-oauth-rfc-deep-dive/",
     "Independent technical blog",
     "Calls out RFC 8707 (Resource Indicators) as the critical compatibility bottleneck since most major IdPs use proprietary audience parameters."),

    ("AI Agents Authentication: How Autonomous Systems Prove Identity (GitGuardian)",
     "A GitGuardian engineering article arguing that delegated, short-lived OAuth grants are structurally safer than the static API keys that produced 28.65M secret leaks in 2025.",
     "https://blog.gitguardian.com/ai-agents-authentication-how-autonomous-systems-prove-identity/",
     "Industry blog (GitGuardian)",
     "Useful for the 'why delegated authorization > shared secrets' business case with concrete 2025 incident data."),

    ("Workload Identity – Key Takeaways from IETF 122 (Defakto)",
     "A practitioner recap of IETF 122 covering the WIMSE Workload Identity Token, Credential Exchange, and side-meeting work on bot/agent web authentication.",
     "https://www.defakto.security/blog/workload-identity-key-takeaways-from-ietf-122/",
     "Industry blog (Defakto)",
     "Author Pieter Kasselman is co-author of multiple key drafts (Transaction Tokens, First-Party Apps); strong primary-source view."),

    ("User-Managed Access (UMA) 2.0 Comprehensive Guide (SSOJet, Feb 2026)",
     "A 2026 industry guide reframing UMA 2.0 for current CIAM contexts, covering resource-set registration, permission tickets, requesting-party tokens (RPTs), and policy externalization gains.",
     "https://ssojet.com/blog/user-managed-access-uma-2-0-comprehensive-guide",
     "Industry blog (SSOJet)",
     "Cites Kantara analysis showing centralized-policy moves can cut authorization code 80% — useful pitch material."),

    ("Decentralized Identity and Verifiable Credentials: The Enterprise Playbook 2026",
     "An enterprise-oriented guide explaining how W3C VCs, DIDs, and OpenID4VC enable the issuer-holder-verifier delegation model for digital credentials and EUDI Wallet acceptance.",
     "https://securityboulevard.com/2026/03/decentralized-identity-and-verifiable-credentials-the-enterprise-playbook-2026/",
     "Industry article (Security Boulevard)",
     "Critical 2027 EU compliance dates: banks, telecom, healthcare, and very-large platforms must accept EUDI Wallet."),

    ("OAuth 2.0 & OpenID Connect: The Complete Guide to What the Standards Actually Say (Patil, Medium)",
     "A practitioner-friendly synthesis covering OAuth 2.0 core, the OAuth 2.1 draft, RFC 9068 JWT access tokens, RFC 8252 native apps, and RFC 8628 device authorization.",
     "https://mrutyunjaypatil.medium.com/oauth-2-0-openid-connect-the-complete-guide-to-what-the-standards-actually-say-e92f040a4251",
     "Industry blog (Medium)",
     "Good handoff document for engineers new to the area; cites the current standards rather than legacy patterns."),

    ("LinkedIn Engineering — OpenID Connect Authentication for Sign In with LinkedIn V2",
     "LinkedIn's own engineering write-up on adopting OpenID Connect on top of OAuth 2.0, explicitly distinguishing OAuth's delegated-access role from OIDC's identity-assertion role.",
     "https://www.linkedin.com/developers/news/featured-updates/openid-connect-authentication",
     "Industry (LinkedIn Engineering)",
     "Real-world example from LinkedIn explaining why they layered OIDC over an existing OAuth 2.0 delegated-access stack."),

    ("Rich Authorization Requests (RAR) — Authlete Knowledge Base",
     "A vendor knowledge-base article giving worked examples of RAR's authorization_details object, including locations, actions, datatypes, and identifier fields with their Authlete bindings.",
     "https://www.authlete.com/kb/oauth-and-openid-connect/authorization-requests/rich-authorization-requests/",
     "Vendor documentation (Authlete)",
     "Authlete is one of the few certified FAPI 2.0 + RAR implementations; useful reference for concrete request bodies."),

    ("An Introduction to Authorization Exchange (AuthZEN) — Curity",
     "A vendor explainer comparing externalized-authorization patterns (OPA, Cedar, XACML, Zanzibar) and showing how AuthZEN's API normalizes the PDP/PEP wire protocol across them.",
     "https://curity.io/resources/learn/authzen/",
     "Vendor blog (Curity)",
     "Curity is a member of the AuthZEN WG; positions AuthZEN as 'OpenID Connect for authorization' which has become the common framing."),

    ("ACLX — AI Output Governance (Proprietary)",
     "A proprietary enforcement layer that intercepts AI-generated content between inference and delivery, evaluating it against OPA/Cedar policies, identity context, and a three-phase sensitivity detection stack: deterministic regex rules → ontology-based compilation-risk scoring → semantic LLM evaluation for novel synthesis. Four enforcement outcomes: ALLOW, REDACT, BLOCK, or ESCALATE to human review. Treats agents as first-class principals evaluated at the output boundary regardless of whether a session or IDP exists.",
     "https://aclx.ai/",
     "Proprietary product",
     "No published specification. Identifies a gap the corpus standards map does not yet cover: identity (OIDC/SCIM/SPIFFE), signals (CAEP/SSF), policy (OAuth/AuthZEN/OPA), and enforcement (ZTA/RATS/SCITT) all have standards, but output evaluation — governing what the AI synthesizes rather than what the user was authorized to ask — has none. The synthesis-detection layer (AI compiling across individually-authorized sources to produce content crossing a higher sensitivity boundary) is a novel problem the corpus's existing standards do not address."),
]
make_sheet("Industry & Implementations", COLORS['Industry'], industry_rows)

# ============================================================
# Index tab
# ============================================================
ws_idx = wb.create_sheet("Index", 0)
fill = PatternFill('solid', start_color=COLORS['Summary'])
ws_idx['A1'] = "Delegated Authorization Research — Index"
ws_idx['A1'].font = Font(name='Arial', size=14, bold=True)
ws_idx.merge_cells('A1:D1')

ws_idx['A2'] = "Research compiled May 2026. Sources separated by maturity so it's clear where active work is happening."
ws_idx['A2'].font = Font(name='Arial', size=10, italic=True, color='595959')
ws_idx.merge_cells('A2:D2')

hdr = ["Tab", "What's in it", "Count", "Where the action is"]
for col, h in enumerate(hdr, 1):
    c = ws_idx.cell(row=4, column=col, value=h)
    c.font, c.fill, c.alignment, c.border = HEADER_FONT, fill, HEADER_ALIGN, BORDER

idx_data = [
    ("Published RFCs", "Settled IETF standards-track and BCP RFCs — the foundation everything else builds on.",
     len(rfc_rows), "Stable. These are the primitives, not where the debate is."),
    ("Active IETF Drafts", "IETF WG charters, requirements drafts, and active WG/individual drafts — including the OAuth WG recharter formally adding 'Complex Delegation' for agents and a six-draft McGuinness individual-submission cluster.",
     len(draft_rows), "★ THIS IS WHERE THE CURRENT WORK IS HAPPENING ★  OAuth recharter on 4 Jun 2026 IESG telechat."),
    ("OpenID Foundation", "Final and draft OIDF specs and the Oct 2025 Agentic AI whitepaper: AuthZEN (incl. the new ARAP profile), Shared Signals/CAEP, FAPI 2.0, HEART.",
     len(oidf_rows), "Mostly Final. AuthZEN Access Request & Approval Profile (ARAP) was adopted as a WG draft May 2026, Draft 1 published 3 Jun 2026."),
    ("Other Standards & Govt", "Kantara UMA 2.0, W3C VCs, NIST AI initiative, EU AI Act compliance dates.",
     len(other_rows), "VC v2.1 First Public Working Draft is the active piece."),
    ("Academic Papers", "arXiv preprints and IEEE conference papers on delegated authz and workload identity.",
     len(academic_rows), "Mostly settled; the 2025 South et al. paper is the most-cited foundation."),
    ("Industry & Implementations", "Vendor blogs, LinkedIn-style articles, reference implementations (OVID/OVID-ME, ZeroID, WorkOS auth.md, Agent Trust Protocol), and the McGuinness Mission-Bound OAuth blog series.",
     len(industry_rows), "Practitioner content; the five implementations at the top of the tab show what an implementable agent-delegation stack looks like today. The four McGuinness Mission-Bound OAuth posts (May–Jun 2026) form a coherent architectural argument."),
]

for i, (tab, what, count, action) in enumerate(idx_data, start=5):
    ws_idx.cell(row=i, column=1, value=tab).font = Font(name='Arial', size=10, bold=True)
    ws_idx.cell(row=i, column=2, value=what).font = BODY_FONT
    ws_idx.cell(row=i, column=3, value=count).font = BODY_FONT
    ws_idx.cell(row=i, column=4, value=action).font = BODY_FONT
    for col in range(1, 5):
        cell = ws_idx.cell(row=i, column=col)
        cell.alignment = BODY_ALIGN
        cell.border = BORDER

total_row = 5 + len(idx_data)
total_count = sum(count for _, _, count, _ in idx_data)
ws_idx.cell(row=total_row, column=1, value="TOTAL").font = Font(name='Arial', size=10, bold=True)
ws_idx.cell(row=total_row, column=3, value=total_count).font = Font(name='Arial', size=10, bold=True)
for col in range(1, 5):
    ws_idx.cell(row=total_row, column=col).border = BORDER

ws_idx.column_dimensions['A'].width = 30
ws_idx.column_dimensions['B'].width = 70
ws_idx.column_dimensions['C'].width = 8
ws_idx.column_dimensions['D'].width = 60
ws_idx.row_dimensions[1].height = 22
ws_idx.row_dimensions[4].height = 30

wb.save('/Users/gffletch/Develop/Authorization/da_research/delegated_authorization_research.xlsx')
print("OK - workbook saved")
print(f"Tab counts: RFCs={len(rfc_rows)}, Drafts={len(draft_rows)}, OIDF={len(oidf_rows)}, Other={len(other_rows)}, Academic={len(academic_rows)}, Industry={len(industry_rows)}")
print(f"Total: {len(rfc_rows)+len(draft_rows)+len(oidf_rows)+len(other_rows)+len(academic_rows)+len(industry_rows)}")
